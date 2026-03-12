from __future__ import annotations

import array
import csv
import io
import json
import math
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

import duckdb
import pandas as pd
from netCDF4 import Dataset
from openpyxl import Workbook

from spectral_library.manifest import SourceRecord
from spectral_library.normalize import (
    SpectrumRecord,
    TARGET_WAVELENGTHS,
    _clean_spectrum,
    _convert_wavelengths_to_nm,
    _detect_value_scale,
    _envi_shape,
    _envi_typecode,
    _interpolate_to_grid,
    _iter_ecostress_text_spectra,
    _iter_envi_spectra,
    _iter_netcdf_spectra,
    _iter_rds_spectra,
    _iter_row_wide_spectra,
    _iter_tabular_spectra,
    _iter_textual_spectra_from_lines,
    _iter_xlsx_spectra,
    _iter_zip_spectra,
    _median_spacing_nm,
    _normalize_header_name,
    _parse_envi_header,
    _parse_envi_scalar,
    _parse_float,
    _parse_wavelength_label,
    _pick_sample_name,
    _resolve_envi_data_path,
    _should_skip_source_path,
    _sniff_delimiter,
    normalize_sources,
)


def manifest_row(**overrides: str) -> dict[str, str]:
    row = {
        "source_id": "src1",
        "name": "Source 1",
        "section": "direct_public",
        "subsection": "soil",
        "spectral_type": "Soil",
        "coverage": "400-2500 nm",
        "resource_type": "Direct library",
        "provider": "manual",
        "landing_url": "https://example.com/src1",
        "download_url": "",
        "fetch_adapter": "manual_portal",
        "auth_mode": "manual_review",
        "expected_format": "csv",
        "tier": "tier1",
        "priority": "high",
        "ingest_role": "primary_raw",
        "normalization_eligibility": "eligible_full",
        "status": "planned",
        "notes": "note",
    }
    row.update(overrides)
    return row


def write_manifest(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(SourceRecord.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_fetch_result(path: Path, source_id: str, status: str = "downloaded") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source_id": source_id,
        "source_name": f"Source {source_id}",
        "fetch_adapter": "manual_portal",
        "fetch_mode": "assets",
        "status": status,
        "landing_url": f"https://example.com/{source_id}",
        "started_at": "2024-01-01T00:00:00+00:00",
        "finished_at": "2024-01-01T00:00:01+00:00",
        "notes": [],
        "artifacts": [],
    }
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


class NormalizeHelpersTests(unittest.TestCase):
    def test_scalar_parsers_and_unit_helpers_cover_edge_cases(self) -> None:
        self.assertIsNone(_parse_float(None))
        self.assertIsNone(_parse_float("bad"))
        self.assertEqual(_parse_float("12%"), 12.0)
        self.assertEqual(_parse_float("0,25"), 0.25)
        self.assertEqual(_normalize_header_name(' "Name" '), "name")

        self.assertEqual(_parse_wavelength_label("X350"), 350.0)
        self.assertEqual(_parse_wavelength_label("SPC.400"), 400.0)
        self.assertEqual(_parse_wavelength_label("450 nm"), 450.0)
        self.assertEqual(_parse_wavelength_label("Rrs_551"), 551.0)
        self.assertEqual(_parse_wavelength_label("Refl (865 nm)"), 865.0)
        self.assertEqual(_parse_wavelength_label("scan_visnir.350_ref"), 350.0)
        self.assertIsNone(_parse_wavelength_label("metadata"))

        self.assertEqual(_detect_value_scale([0.2, 0.5]), 1.0)
        self.assertEqual(_detect_value_scale([20.0, 40.0]), 100.0)
        self.assertEqual(_detect_value_scale([200.0, 400.0]), 1000.0)
        self.assertEqual(_detect_value_scale([2000.0, 4000.0]), 10000.0)
        self.assertEqual(_detect_value_scale([425.0, 14415.0]), 10000.0)
        self.assertEqual(_detect_value_scale([], hint=5.0), 5.0)

        self.assertEqual(_convert_wavelengths_to_nm([], "nm"), [])
        self.assertEqual(_convert_wavelengths_to_nm([0.4, 0.5], "micrometers"), [400.0, 500.0])
        self.assertEqual(_convert_wavelengths_to_nm([0.4, 0.5], ""), [400.0, 500.0])
        self.assertEqual(_convert_wavelengths_to_nm([400.0, 500.0], "nm"), [400.0, 500.0])
        self.assertEqual(_convert_wavelengths_to_nm([3.5e-7, 3.51e-7], "m"), [350.0, 351.0])

        self.assertEqual(_median_spacing_nm([400.0, 401.0, 403.0]), 1.5)
        self.assertIsNone(_median_spacing_nm([400.0, 400.0]))

    def test_clean_spectrum_handles_duplicate_and_invalid_pairs(self) -> None:
        dedup_record = SpectrumRecord(
            source_id="src1",
            source_name="Source 1",
            ingest_role="primary_raw",
            input_path="memory",
            parser="test",
            sample_name="sample_a",
            wavelengths_nm=[401.0, 400.0, 400.0, 402.0, 403.0],
            values=[0.2, 0.1, 0.15, math.nan, 0.4],
            metadata={},
        )

        wavelengths, values, scale = _clean_spectrum(dedup_record)
        self.assertEqual(scale, 1.0)
        self.assertEqual(wavelengths, [400.0, 401.0, 403.0])
        self.assertEqual(values, [0.15, 0.2, 0.4])

        bad_record = SpectrumRecord(
            source_id="src1",
            source_name="Source 1",
            ingest_role="primary_raw",
            input_path="memory",
            parser="test",
            sample_name="sample_b",
            wavelengths_nm=[400.0],
            values=[0.1],
            metadata={},
        )
        with self.assertRaises(ValueError):
            _clean_spectrum(bad_record)

    def test_envi_helpers_cover_multiline_and_error_branches(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            multiline_header = root / "multiline.hdr"
            multiline_header.write_text(
                "\n".join(
                    [
                        "ENVI",
                        "description = {line one,",
                        " line two}",
                        "ignored line",
                        "data file = explicit.img",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (root / "explicit.img").write_bytes(array.array("f", [1.0, 2.0]).tobytes())

            parsed = _parse_envi_header(multiline_header)
            self.assertIn("description", parsed)
            self.assertEqual(_parse_envi_scalar("{}"), [])
            self.assertEqual(_resolve_envi_data_path(multiline_header, parsed), root / "explicit.img")

            image_header = root / "image_only.hdr"
            image_header.write_text("ENVI\nsamples = 1\n", encoding="utf-8")
            (root / "image_only.img").write_bytes(b"xx")
            self.assertEqual(_resolve_envi_data_path(image_header, {}), root / "image_only.img")

            with self.assertRaises(FileNotFoundError):
                _resolve_envi_data_path(root / "missing.hdr", {})

        self.assertEqual(_envi_shape({"samples": "4", "lines": "3", "bands": "1", "file type": "ENVI Spectral Library"}), (3, 4))
        self.assertEqual(_envi_shape({"samples": "1", "lines": "2", "bands": "5", "interleave": "bip"}), (2, 5))
        self.assertEqual(_envi_shape({"samples": "4", "lines": "2", "bands": "1", "interleave": "bsq"}), (2, 4))
        with self.assertRaises(ValueError):
            _envi_shape({"samples": "2", "lines": "2", "bands": "2", "interleave": "bil"})
        with self.assertRaises(ValueError):
            _envi_typecode(99)

    def test_sample_name_and_delimiter_helpers_cover_fallbacks(self) -> None:
        self.assertEqual(_pick_sample_name({"Name": "oak"}, 1), "oak")
        self.assertEqual(_pick_sample_name({"Name": "NA"}, 3), "spectrum_00003")
        self.assertEqual(_sniff_delimiter("A;B;C\n1;2;3\n"), ";")
        self.assertEqual(_sniff_delimiter("only text without csv structure"), ",")

    def test_clean_spectrum_scales_percent_values_and_interpolates(self) -> None:
        record = SpectrumRecord(
            source_id="src1",
            source_name="Source 1",
            ingest_role="primary_raw",
            input_path="memory",
            parser="test",
            sample_name="sample_a",
            wavelengths_nm=[400.0, 500.0, 600.0],
            values=[20.0, 40.0, 60.0],
            metadata={},
        )

        wavelengths, values, scale = _clean_spectrum(record)
        normalized = _interpolate_to_grid(wavelengths, values)

        self.assertEqual(scale, 100.0)
        self.assertAlmostEqual(values[0], 0.2)
        self.assertAlmostEqual(normalized[0], 0.2)
        self.assertAlmostEqual(normalized[100], 0.4)
        self.assertAlmostEqual(normalized[200], 0.6)

    def test_iter_envi_spectra_reads_spectral_library_pairs(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            header_path = root / "library.hdr"
            data_path = root / "library.sli"
            header_path.write_text(
                "\n".join(
                    [
                        "ENVI",
                        "samples = 3",
                        "lines = 2",
                        "bands = 1",
                        "file type = ENVI Spectral Library",
                        "data type = 4",
                        "interleave = bsq",
                        "byte order = 0",
                        "wavelength units = Nanometers",
                        "wavelength = { 400, 401, 402 }",
                        "spectra names = { alpha, beta }",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            values = array.array("f", [0.1, 0.2, 0.3, 0.4, 0.5, 0.6])
            data_path.write_bytes(values.tobytes())

            spectra = list(_iter_envi_spectra(header_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(spectra), 2)
            self.assertEqual(spectra[0].sample_name, "alpha")
            self.assertEqual(spectra[1].sample_name, "beta")
            self.assertEqual(spectra[0].wavelengths_nm, [400.0, 401.0, 402.0])
            self.assertAlmostEqual(spectra[1].values[-1], 0.6)

    def test_iter_tabular_spectra_handles_row_column_and_long_layouts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)

            row_wide = root / "row_wide.csv"
            row_wide.write_text(
                "sample_name,400,401,402\nsample_a,0.1,0.2,0.3\n",
                encoding="utf-8",
            )
            row_records = list(_iter_tabular_spectra(row_wide, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(row_records), 1)
            self.assertEqual(row_records[0].parser, "csv_row_wide")
            self.assertEqual(row_records[0].sample_name, "sample_a")

            column_wide = root / "column_wide.csv"
            column_wide.write_text(
                "wavelength,A,B\n400,0.1,0.3\n401,0.2,0.4\n",
                encoding="utf-8",
            )
            column_records = list(_iter_tabular_spectra(column_wide, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(column_records), 2)
            self.assertEqual(column_records[0].parser, "csv_column_wide")
            self.assertEqual(column_records[1].sample_name, "B")

            long_table = root / "long_table.txt"
            long_table.write_text(
                "\n".join(
                    [
                        "# comment",
                        "record,wavelength,R,observation_ID",
                        "1,400,0.1,obs_1",
                        "2,401,0.2,obs_1",
                        "3,400,0.3,obs_2",
                        "4,401,0.4,obs_2",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            long_records = list(_iter_tabular_spectra(long_table, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(long_records), 2)
            self.assertEqual(long_records[0].parser, "csv_long_table")
            self.assertEqual(long_records[0].sample_name, "obs_1")

            neon_long = root / "neon_long.csv"
            neon_long.write_text(
                "\n".join(
                    [
                        "spectralSampleID,spectralSampleCode,wavelength,reflectanceCondition,reflectance",
                        "FSP_DSNY_20140508_1449,,350,top of foliage,0.11",
                        "FSP_DSNY_20140508_1449,,351,top of foliage,0.12",
                        "FSP_DSNY_20140508_1454,,350,top of foliage,0.21",
                        "FSP_DSNY_20140508_1454,,351,top of foliage,0.22",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            neon_records = list(_iter_tabular_spectra(neon_long, "neon_field_spectra", "NEON", "primary_raw"))
            self.assertEqual(len(neon_records), 2)
            self.assertEqual(neon_records[0].sample_name, "FSP_DSNY_20140508_1449")

            empty = root / "empty.csv"
            empty.write_text("", encoding="utf-8")
            self.assertEqual(list(_iter_tabular_spectra(empty, "src1", "Source 1", "primary_raw")), [])

    def test_iter_tabular_spectra_covers_additional_branch_behaviors(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)

            blank_column_file = root / "blank_column.csv"
            blank_column_file.write_text("wavelength\n400\n", encoding="utf-8")
            self.assertEqual(list(_iter_tabular_spectra(blank_column_file, "src1", "Source 1", "primary_raw")), [])

            long_missing = root / "long_missing.csv"
            long_missing.write_text(
                "record,wavelength,R,id\n1,400,0.1,\n2,401,0.2,\n3,402,,\n",
                encoding="utf-8",
            )
            records = list(_iter_tabular_spectra(long_missing, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(records), 2)
            self.assertEqual(records[0].sample_name, "spectrum_00001")

    def test_iter_tabular_spectra_handles_comment_blocks_and_two_row_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)

            block_comment = root / "commented.tab"
            block_comment.write_text(
                "\n".join(
                    [
                        "/* metadata block",
                        "ignored line",
                        "*/",
                        "Event\tRefl (350 nm)\tRefl (351 nm)\tRefl (352 nm)",
                        "sample_a\t0.1\t0.2\t0.3",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            comment_records = list(_iter_tabular_spectra(block_comment, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(comment_records), 1)
            self.assertEqual(comment_records[0].parser, "csv_row_wide")

            two_row = root / "two_row.csv"
            two_row.write_text(
                "\n".join(
                    [
                        "index,spectra,,,",
                        "lambda,350,351,352",
                        "A001,0.1,0.2,0.3",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            two_row_records = list(_iter_tabular_spectra(two_row, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(two_row_records), 1)
            self.assertEqual(two_row_records[0].sample_name, "A001")
            self.assertEqual(two_row_records[0].parser, "csv_row_wide")

            generic_id = Path(tmpdir) / "generic_id.csv"
            generic_id.write_text(
                "GLORIA_ID,Es_350,Es_351,Es_352\nGID_1,0.1,0.2,0.3\n",
                encoding="utf-8",
            )
            generic_records = list(_iter_tabular_spectra(generic_id, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(generic_records), 1)
            self.assertEqual(generic_records[0].sample_name, "GID_1")

    def test_iter_ecostress_text_spectra_converts_units_and_percent(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.txt"
            path.write_text(
                "\n".join(
                    [
                        "Name: Distilled Water",
                        "X Units: Wavelength (micrometers)",
                        "Y Units: Reflectance (percent)",
                        "",
                        "0.400 10.0",
                        "0.500 20.0",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            records = list(_iter_ecostress_text_spectra(path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].sample_name, "Distilled Water")
            self.assertEqual(records[0].wavelengths_nm, [400.0, 500.0])
            self.assertEqual(records[0].value_scale_hint, 100.0)

    def test_iter_envi_spectra_raises_for_missing_wavelengths_and_bad_lengths(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            bad_header = root / "bad.hdr"
            bad_header.write_text(
                "\n".join(
                    [
                        "ENVI",
                        "samples = 2",
                        "lines = 1",
                        "bands = 1",
                        "file type = ENVI Spectral Library",
                        "data type = 4",
                        "interleave = bsq",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (root / "bad.sli").write_bytes(array.array("f", [1.0, 2.0]).tobytes())
            with self.assertRaises(ValueError):
                list(_iter_envi_spectra(bad_header, "src1", "Source 1", "primary_raw"))

            mismatch_header = root / "mismatch.hdr"
            mismatch_header.write_text(
                "\n".join(
                    [
                        "ENVI",
                        "samples = 3",
                        "lines = 1",
                        "bands = 1",
                        "file type = ENVI Spectral Library",
                        "data type = 4",
                        "interleave = bsq",
                        "wavelength = { 400, 401, 402 }",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (root / "mismatch.sli").write_bytes(array.array("f", [1.0, 2.0]).tobytes())
            with self.assertRaises(ValueError):
                list(_iter_envi_spectra(mismatch_header, "src1", "Source 1", "primary_raw"))

    def test_iter_textual_spectra_and_zip_cover_usgs_and_archive_members(self) -> None:
        usgs_lines = ["s07_ASD Record=1: Example"] + ["-1.23e34"] + ["0.1"] * 1000 + ["-1.23e34"]
        usgs_records = list(
            _iter_textual_spectra_from_lines(usgs_lines, "memory::sample.txt", "src1", "Source 1", "primary_raw")
        )
        self.assertEqual(len(usgs_records), 1)
        self.assertEqual(usgs_records[0].parser, "usgs_ascii")
        self.assertEqual(usgs_records[0].wavelengths_nm[0], 350)

        with tempfile.TemporaryDirectory() as tmpdir:
            archive_path = Path(tmpdir) / "bundle.zip"
            with zipfile.ZipFile(archive_path, "w") as archive:
                archive.writestr(
                    "sample_metadata.csv",
                    "index,class\nA001,roof\n",
                )
                archive.writestr(
                    "sample_quality.csv",
                    "index,quality,889,890\nA001,10,10,10\n",
                )
                archive.writestr(
                    "spectra.csv",
                    "Wavelength (nm),sample_a,sample_b\n400,0.1,0.3\n401,0.2,0.4\n",
                )
            records = list(_iter_zip_spectra(archive_path, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(records), 2)
            self.assertEqual(records[0].parser, "csv_column_wide")

    def test_iter_zip_spectra_skips_usgs_auxiliary_members_and_uses_wavelength_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            archive_path = Path(tmpdir) / "usgs_bundle.zip"
            spectrum_lines = ["s07_ASD Record=1: Example spectrum"] + ["0.1"] * 1001
            wavelength_lines = ["s07_ASD Record=19: Wavelengths ASD 0.35-2.5 microns 2151 ch"] + [
                f"{0.35 + 0.001 * idx:.6f}" for idx in range(1001)
            ]
            errorbar_lines = ["s07_ASD Record=2: Example error bars"] + ["0.01"] * 1001
            with zipfile.ZipFile(archive_path, "w") as archive:
                archive.writestr(
                    "ASCIIdata_splib07b_cvASD/ChapterA/s07_ASD_example.txt",
                    "\n".join(spectrum_lines),
                )
                archive.writestr(
                    "ASCIIdata_splib07b_cvASD/errorbars/errorbars_for_s07_ASD_example.txt",
                    "\n".join(errorbar_lines),
                )
                archive.writestr(
                    "ASCIIdata_splib07b_cvASD/s07_ASD_Wavelengths_ASD_0.35-2.5_microns_2151_ch.txt",
                    "\n".join(wavelength_lines),
                )
                archive.writestr(
                    "ASCIIdata_splib07b_cvASD/s07_ASD_Bandpass_(FWHM)_ASDFR_StandardResolution.txt",
                    "\n".join(errorbar_lines),
                )

            records = list(_iter_zip_spectra(archive_path, "usgs_v7", "USGS v7", "primary_raw"))
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].parser, "usgs_ascii")
            self.assertEqual(records[0].wavelengths_nm[:3], [350.0, 351.0, 352.0])
            self.assertEqual(records[0].sample_name, "Example spectrum")

    def test_should_skip_source_path_skips_extracted_usgs_auxiliary_files(self) -> None:
        self.assertTrue(
            _should_skip_source_path(
                "usgs_v7",
                Path("/tmp/usgs/data/ASCIIdata_splib07b_cvASD/errorbars/errorbars_for_s07_ASD_example.txt"),
            )
        )
        self.assertTrue(
            _should_skip_source_path(
                "usgs_v7",
                Path("/tmp/usgs/data/ASCIIdata_splib07b_cvASD/s07_ASD_Wavelengths_ASD_0.35-2.5_microns_2151_ch.txt"),
            )
        )
        self.assertFalse(
            _should_skip_source_path(
                "usgs_v7",
                Path("/tmp/usgs/data/ASCIIdata_splib07b_cvASD/ChapterV_Vegetation/s07_ASD_Willow_leaf.txt"),
            )
        )

    def test_iter_zip_spectra_recurses_into_nested_archives(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            archive_path = Path(tmpdir) / "outer.zip"
            inner_buffer = io.BytesIO()
            with zipfile.ZipFile(inner_buffer, "w") as inner_archive:
                inner_archive.writestr(
                    "SPECTRAL LIBRARY/replicate sample means.csv",
                    "Wavelength (nm),sample_a,sample_b\n400,0.1,0.3\n401,0.2,0.4\n",
                )
                inner_archive.writestr(
                    "SPECTRAL LIBRARY/information.txt",
                    "metadata only\n",
                )
            with zipfile.ZipFile(archive_path, "w") as archive:
                archive.writestr("README.txt", "outer readme\n")
                archive.writestr("nested_library.zip", inner_buffer.getvalue())

            records = list(_iter_zip_spectra(archive_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(records), 2)
            self.assertTrue(all(record.parser == "csv_column_wide" for record in records))
            self.assertTrue(all("nested_library.zip" in record.input_path for record in records))

    def test_iter_zip_spectra_skips_metadata_csv_members(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            archive_path = Path(tmpdir) / "bundle.zip"
            with zipfile.ZipFile(archive_path, "w") as archive:
                archive.writestr(
                    "Processed/Spectra.metadata.csv",
                    "sample_id,note\nA001,test\n",
                )
                archive.writestr(
                    "Processed/Leaf_Reflectance.csv",
                    "sample_id,400,401,402\nleaf_a,0.1,0.2,0.3\n",
                )

            records = list(_iter_zip_spectra(archive_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].sample_name, "leaf_a")

    def test_iter_xlsx_spectra_handles_row_wide_and_band_matrix_layouts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "spectra.xlsx"
            workbook = Workbook()
            wide = workbook.active
            wide.title = "Wide"
            wide.append(["ID_Unique", 350, 351, 352])
            wide.append(["sample_a", 0.1, 0.2, 0.3])

            band = workbook.create_sheet("BandMatrix")
            band.append(["Title"])
            band.append(["Julian Day", None, 126, 127, 128])
            band.append(["Plot ID", None, "A1", "A2", "A3"])
            band.append([None, "band", "Reflectance", "Reflectance", "Reflectance"])
            band.append([None, 351, 0.11, 0.21, 0.31])
            band.append([None, 361, 0.12, 0.22, 0.32])
            workbook.save(workbook_path)

            records = list(_iter_xlsx_spectra(workbook_path, "src1", "Source 1", "primary_raw"))
            parsers = [record.parser for record in records]
            self.assertIn("xlsx_row_wide", parsers)
            self.assertIn("xlsx_band_matrix", parsers)
            self.assertTrue(any(record.sample_name == "sample_a" for record in records))
            self.assertTrue(any(record.sample_name.startswith("A1_") for record in records))

    def test_iter_xlsx_spectra_forces_unit_scale_for_understory_estonia(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "understory.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Spectra"
            sheet.append(["Stand nr", "Position", "WL350", "WL351", "WL352"])
            sheet.append([1, 1, 0.02, 0.03, 500.0])
            workbook.save(workbook_path)

            records = list(_iter_xlsx_spectra(workbook_path, "understory_estonia_czech", "Understory", "primary_raw"))

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].parser, "xlsx_row_wide")
            self.assertEqual(records[0].value_scale_hint, 1.0)

    def test_iter_row_wide_spectra_forces_percent_scale_for_verified_sources(self) -> None:
        cases = [
            ("ossl", ["id.scan_uuid_c", "scan_visnir.350_pcnt", "scan_visnir.352_pcnt"], {"id.scan_uuid_c": "sample", "scan_visnir.350_pcnt": 4.7, "scan_visnir.352_pcnt": 4.9}),
            ("ghisacasia_v001", ["Spectra", "350", "351"], {"Spectra": "C1A_jd152", "350": 23.5, "351": 24.1}),
            (
                "ngee_arctic_leaf_reflectance_transmittance_barrow_2014_2016",
                ["Spectra_Name", "Wave_350", "Wave_351"],
                {"Spectra_Name": "PEFR5_1624_IS_T", "Wave_350": 12.0, "Wave_351": -0.33},
            ),
        ]

        for source_id, header, row in cases:
            with self.subTest(source_id=source_id):
                records = list(
                    _iter_row_wide_spectra(
                        "memory.csv",
                        source_id,
                        source_id,
                        "primary_raw",
                        header,
                        [row],
                    )
                )
                self.assertEqual(len(records), 1)
                self.assertEqual(records[0].value_scale_hint, 100.0)

    def test_iter_row_wide_spectra_uses_hyspiri_row_specific_scale_hints(self) -> None:
        ratio_header = ["Spectra", *[str(350 + index) for index in range(6)]]
        ratio_row = {
            "Spectra": "Ivanpah_1",
            **{str(350 + index): value for index, value in enumerate([0.15, 0.4, 0.51, 0.56, 0.49, 0.52])},
        }
        percent_row = {
            "Spectra": "Bare_Field_1_Spec-00841",
            **{str(350 + index): value for index, value in enumerate([3.1, 25.0, 33.2, 54.8, 104.7, 45.2])},
        }
        spectralon_header = ["Spectra", *[str(350 + index) for index in range(100)]]
        spectralon_values = [99.8] * 99 + [239.283]
        spectralon_row = {
            "Spectra": "Bare_Field_1_Spectralon",
            **{str(350 + index): value for index, value in enumerate(spectralon_values)},
        }

        ratio_record = list(
            _iter_row_wide_spectra(
                "uw-bnl_nasa_hyspiri_airborne_campaign_ground_cal_target_spectra_spectral_measurements.csv",
                "hyspiri_ground_targets",
                "HyspIRI",
                "primary_raw",
                ratio_header,
                [ratio_row],
            )
        )[0]
        percent_record = list(
            _iter_row_wide_spectra(
                "uw-bnl_nasa_hyspiri_airborne_campaign_ground_cal_target_spectra_spectral_measurements.csv",
                "hyspiri_ground_targets",
                "HyspIRI",
                "primary_raw",
                ratio_header,
                [percent_row],
            )
        )[0]
        spectralon_record = list(
            _iter_row_wide_spectra(
                "uw-bnl_nasa_hyspiri_airborne_campaign_ground_cal_target_spectra_spectral_measurements.csv",
                "hyspiri_ground_targets",
                "HyspIRI",
                "primary_raw",
                spectralon_header,
                [spectralon_row],
            )
        )[0]

        self.assertEqual(ratio_record.value_scale_hint, 1.0)
        self.assertEqual(percent_record.value_scale_hint, 100.0)
        self.assertEqual(spectralon_record.value_scale_hint, 100.0)

    def test_iter_textual_spectra_for_hyspiri_ivanpah_uses_only_reflectance_column(self) -> None:
        lines = [
            "Measurement: REFLECTANCE,,,",
            "Data:,,,",
            "Wvl,Norm. DN (Ref.),Norm. DN (Target),Reflect. %",
            "350,723,112,15.977",
            "351,741,115,15.990",
            "352,765,119,16.037",
        ]

        records = list(
            _iter_textual_spectra_from_lines(
                lines,
                "ivanpah_dry_lake_spectra_20130603.zip::Playa1/Ivanpah1-Spec_00751.csv",
                "hyspiri_ground_targets",
                "HyspIRI",
                "primary_raw",
            )
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].value_scale_hint, 100.0)
        self.assertEqual(records[0].values[:3], [15.977, 15.99, 16.037])

    def test_iter_xlsx_spectra_handles_mean_curve_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "means.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "FI-Hyy 20180628"
            sheet.append(["wavelength", "mean", "standev"])
            sheet.append([400, 0.11, 0.01])
            sheet.append([401, 0.12, 0.02])
            sheet.append([402, 0.13, 0.03])
            workbook.save(workbook_path)

            records = list(_iter_xlsx_spectra(workbook_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].parser, "xlsx_sheet_mean_curve")
            self.assertEqual(records[0].sample_name, "FI-Hyy 20180628")
            self.assertEqual(records[0].wavelengths_nm, [400.0, 401.0, 402.0])
            self.assertEqual(records[0].values, [0.11, 0.12, 0.13])

    def test_iter_row_wide_spectra_filters_auxiliary_stat_band_fields(self) -> None:
        header = ["sample_id", "reflectance_400", "reflectance_401", "std_400", "error_401", "uncertainty_402"]
        rows = [
            {
                "sample_id": "sample_a",
                "reflectance_400": 0.1,
                "reflectance_401": 0.2,
                "std_400": 0.01,
                "error_401": 0.02,
                "uncertainty_402": 0.03,
            }
        ]

        records = list(
            _iter_row_wide_spectra(
                "sample.csv",
                "src1",
                "Source 1",
                "primary_raw",
                header,
                rows,
            )
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].wavelengths_nm, [400.0, 401.0])
        self.assertEqual(records[0].values, [0.1, 0.2])

    def test_iter_tabular_spectra_filters_auxiliary_column_series(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "columns.csv"
            csv_path.write_text(
                "wavelength,mean,std,stdv,stderr,error,uncertainty,SampleA\n"
                "400,0.1,0.01,0.011,0.012,0.013,0.014,0.5\n"
                "401,0.2,0.02,0.021,0.022,0.023,0.024,0.6\n",
                encoding="utf-8",
            )

            records = list(_iter_tabular_spectra(csv_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual([record.sample_name for record in records], ["mean", "SampleA"])
            self.assertEqual(records[0].values, [0.1, 0.2])
            self.assertEqual(records[1].values, [0.5, 0.6])

    def test_iter_netcdf_and_rds_spectra_cover_new_binary_readers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            nc_path = root / "reflectance.nc"
            dataset = Dataset(nc_path, "w")
            try:
                dataset.createDimension("obs", 2)
                dataset.createDimension("wavelength", 3)
                wavelength = dataset.createVariable("wavelength", "f4", ("wavelength",))
                wavelength.units = "m"
                wavelength[:] = [3.5e-7, 3.51e-7, 3.52e-7]
                reflectance = dataset.createVariable("reflectance", "f4", ("obs", "wavelength"))
                reflectance[:] = [[0.1, 0.2, 0.3], [0.4, 0.5, 0.6]]
            finally:
                dataset.close()

            netcdf_records = list(_iter_netcdf_spectra(nc_path, "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(netcdf_records), 2)
            self.assertEqual(netcdf_records[0].parser, "netcdf_reflectance")
            for actual, expected in zip(netcdf_records[0].wavelengths_nm, [350.0, 351.0, 352.0]):
                self.assertAlmostEqual(actual, expected, places=4)

            dataframe = pd.DataFrame({"id_unique": ["sample_a"], 350: [0.1], 351: [0.2], 352: [0.3]})
            with mock.patch("rdata.read_rds", return_value=dataframe):
                rds_records = list(_iter_rds_spectra(root / "sample.rds", "src1", "Source 1", "primary_raw"))
            self.assertEqual(len(rds_records), 1)
            self.assertEqual(rds_records[0].parser, "rds_dataframe")
            self.assertEqual(rds_records[0].sample_name, "sample_a")

    def test_iter_netcdf_spectra_ignores_uncertainty_variables(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            nc_path = Path(tmpdir) / "surface.nc"
            dataset = Dataset(nc_path, "w")
            try:
                dataset.createDimension("obs", 1)
                dataset.createDimension("wavelength", 3)
                wavelength = dataset.createVariable("wavelength", "f4", ("wavelength",))
                wavelength.units = "nm"
                wavelength[:] = [400, 401, 402]
                uncertainty = dataset.createVariable("surface_reflectance_uncertainty", "f4", ("obs", "wavelength"))
                uncertainty[:] = [[9.0, 9.0, 9.0]]
                reflectance = dataset.createVariable("surface_reflectance", "f4", ("obs", "wavelength"))
                reflectance[:] = [[0.1, 0.2, 0.3]]
            finally:
                dataset.close()

            records = list(_iter_netcdf_spectra(nc_path, "src1", "Source 1", "primary_raw"))

            self.assertEqual(len(records), 1)
            for actual, expected in zip(records[0].values, [0.1, 0.2, 0.3]):
                self.assertAlmostEqual(actual, expected, places=6)


class NormalizePipelineTests(unittest.TestCase):
    def test_normalize_sources_writes_database_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(
                manifest_path,
                [
                    manifest_row(source_id="src1", name="Source 1"),
                    manifest_row(source_id="src2", name="Source 2", status="review_required"),
                ],
            )
            write_fetch_result(results_root / "src1" / "fetch-result.json", "src1")
            write_fetch_result(results_root / "src2" / "fetch-result.json", "src2", status="metadata_only")

            data_dir = results_root / "src1" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
            (data_dir / "wide.csv").write_text(
                "sample_name,400,401,402\nsample_a,10,20,30\nsample_b,40,50,60\n",
                encoding="utf-8",
            )
            (data_dir / "notes.pdf").write_text("not a spectral file", encoding="utf-8")

            summary = normalize_sources(manifest_path, results_root, output_root)

            self.assertEqual(summary["selected_sources"], 2)
            self.assertEqual(summary["downloaded_sources"], 1)
            self.assertEqual(summary["normalized_sources"], 1)
            self.assertEqual(summary["normalized_spectra"], 2)
            self.assertGreaterEqual(summary["failure_rows"], 1)
            self.assertTrue((output_root / "db" / "normalized_catalog.duckdb").exists())
            self.assertTrue((output_root / "parquet" / "normalized_spectra.parquet").exists())

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                spectra_count = connection.execute("SELECT COUNT(*) FROM normalized_spectra").fetchone()[0]
                metadata_count = connection.execute("SELECT COUNT(*) FROM spectra_metadata").fetchone()[0]
                failure_count = connection.execute("SELECT COUNT(*) FROM normalization_failures").fetchone()[0]
                first_row = connection.execute(
                    "SELECT source_id, spectrum_id, sample_name, nm_400, nm_401, nm_402 FROM normalized_spectra ORDER BY spectrum_id LIMIT 1"
                ).fetchone()
            finally:
                connection.close()

            self.assertEqual(spectra_count, 2)
            self.assertEqual(metadata_count, 2)
            self.assertGreaterEqual(failure_count, 1)
            self.assertEqual(first_row[0], "src1")
            self.assertAlmostEqual(first_row[3], 0.1)
            self.assertAlmostEqual(first_row[4], 0.2)
            self.assertAlmostEqual(first_row[5], 0.3)

            summary_payload = json.loads((output_root / "build_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary_payload["normalized_spectra"], 2)
            self.assertEqual(summary_payload["target_start_nm"], TARGET_WAVELENGTHS[0])

    def test_normalize_sources_scans_root_level_archives_for_usgs_style_sources(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(manifest_path, [manifest_row(source_id="usgs_v7", name="USGS v7")])
            write_fetch_result(results_root / "usgs_v7" / "fetch-result.json", "usgs_v7")

            archive_path = results_root / "usgs_v7" / "ASCIIdata_splib07b_cvASD.zip"
            with zipfile.ZipFile(archive_path, "w") as archive:
                values = ["-1.23e34"] + ["0.1"] * 1000 + ["-1.23e34"]
                archive.writestr(
                    "ASCIIdata_splib07b_cvASD/example.txt",
                    "\n".join(["s07_ASD Record=1: Example", *values]) + "\n",
                )

            summary = normalize_sources(manifest_path, results_root, output_root)

            self.assertEqual(summary["downloaded_sources"], 1)
            self.assertEqual(summary["normalized_sources"], 1)
            self.assertEqual(summary["parser_counts"]["usgs_ascii"], 1)

    def test_normalize_sources_covers_mixed_source_status_and_failure_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(
                manifest_path,
                [
                    manifest_row(source_id="missing_fetch", name="Missing Fetch"),
                    manifest_row(source_id="metadata_only", name="Metadata Only"),
                    manifest_row(source_id="missing_data", name="Missing Data"),
                    manifest_row(source_id="txt_src", name="TXT Source"),
                    manifest_row(source_id="envi_src", name="ENVI Source"),
                    manifest_row(source_id="bad_hdr", name="Bad Header"),
                    manifest_row(source_id="bad_norm", name="Bad Normalize"),
                ],
            )

            write_fetch_result(results_root / "metadata_only" / "fetch-result.json", "metadata_only", status="metadata_only")
            write_fetch_result(results_root / "missing_data" / "fetch-result.json", "missing_data")
            write_fetch_result(results_root / "txt_src" / "fetch-result.json", "txt_src")
            write_fetch_result(results_root / "envi_src" / "fetch-result.json", "envi_src")
            write_fetch_result(results_root / "bad_hdr" / "fetch-result.json", "bad_hdr")
            write_fetch_result(results_root / "bad_norm" / "fetch-result.json", "bad_norm")

            txt_data = results_root / "txt_src" / "data"
            txt_data.mkdir(parents=True, exist_ok=True)
            (txt_data / "sample.txt").write_text(
                "\n".join(
                    [
                        "Name: Soil",
                        "X Units: Wavelength (micrometers)",
                        "Y Units: Reflectance (percent)",
                        "",
                        "0.400 10",
                        "0.500 20",
                        "0.600 30",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (txt_data / "archive.zip").write_text("not supported", encoding="utf-8")
            (txt_data / "extra.bin").write_text("not supported", encoding="utf-8")

            envi_data = results_root / "envi_src" / "data"
            envi_data.mkdir(parents=True, exist_ok=True)
            envi_header = envi_data / "library.hdr"
            envi_header.write_text(
                "\n".join(
                    [
                        "ENVI",
                        "samples = 2",
                        "lines = 1",
                        "bands = 1",
                        "file type = ENVI Spectral Library",
                        "data type = 4",
                        "interleave = bsq",
                        "wavelength units = Nanometers",
                        "wavelength = { 400, 401 }",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (envi_data / "library.sli").write_bytes(array.array("f", [0.1, 0.2]).tobytes())

            bad_hdr_data = results_root / "bad_hdr" / "data"
            bad_hdr_data.mkdir(parents=True, exist_ok=True)
            (bad_hdr_data / "bad.hdr").write_text(
                "\n".join(
                    [
                        "ENVI",
                        "samples = 2",
                        "lines = 1",
                        "bands = 1",
                        "file type = ENVI Spectral Library",
                        "data type = 4",
                        "interleave = bsq",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (bad_hdr_data / "bad.sli").write_bytes(array.array("f", [1.0, 2.0]).tobytes())

            bad_norm_data = results_root / "bad_norm" / "data"
            bad_norm_data.mkdir(parents=True, exist_ok=True)
            (bad_norm_data / "single_point.csv").write_text("wavelength,A\n400,0.1\n", encoding="utf-8")

            summary = normalize_sources(
                manifest_path,
                results_root,
                output_root,
                source_ids=["missing_data", "txt_src", "envi_src", "bad_hdr", "bad_norm", "metadata_only", "missing_fetch"],
            )

            self.assertEqual(summary["selected_sources"], 7)
            self.assertEqual(summary["downloaded_sources"], 5)
            self.assertEqual(summary["normalized_sources"], 2)
            self.assertEqual(summary["parser_counts"]["ecostress_txt"], 1)
            self.assertEqual(summary["parser_counts"]["envi_binary"], 1)
            self.assertGreaterEqual(summary["failure_rows"], 5)

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                failure_reasons = connection.execute(
                    "SELECT parser, stage, reason FROM normalization_failures ORDER BY parser, stage"
                ).fetchall()
                source_rows = connection.execute(
                    "SELECT source_id, normalized_spectra, failure_count FROM source_summary ORDER BY source_id"
                ).fetchall()
            finally:
                connection.close()

            self.assertIn(("source_scan", "scan", "missing data directory"), failure_reasons)
            self.assertTrue(any(row[0] == "txt_src" and row[1] == 1 for row in source_rows))
            self.assertTrue(any(row[0] == "bad_norm" and row[2] >= 1 for row in source_rows))

            limited_output = root / "normalized_limited"
            limited_summary = normalize_sources(manifest_path, results_root, limited_output, source_ids=["txt_src", "envi_src"], limit=1)
            self.assertEqual(limited_summary["selected_sources"], 1)

    def test_normalize_sources_prefers_cabo_reflectance_file_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(manifest_path, [manifest_row(source_id="cabo_leaf_v2", name="CABO Leaf v2")])
            write_fetch_result(results_root / "cabo_leaf_v2" / "fetch-result.json", "cabo_leaf_v2")

            data_dir = results_root / "cabo_leaf_v2" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
            (data_dir / "ref_spec.csv").write_text(
                "sample_name,400,401,402\nleaf_a,0.1,0.2,0.3\n",
                encoding="utf-8",
            )
            (data_dir / "abs_spec.csv").write_text(
                "sample_name,400,401,402\nleaf_a,0.7,0.8,0.9\n",
                encoding="utf-8",
            )
            (data_dir / "trans_spec.csv").write_text(
                "sample_name,400,401,402\nleaf_a,0.01,0.02,0.03\n",
                encoding="utf-8",
            )
            (data_dir / "metadata_fields.csv").write_text("field,description\nx,y\n", encoding="utf-8")

            summary = normalize_sources(manifest_path, results_root, output_root)
            self.assertEqual(summary["normalized_sources"], 1)
            self.assertEqual(summary["normalized_spectra"], 1)

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                row = connection.execute(
                    """
                    SELECT s.sample_name, m.parser, s.nm_400, s.nm_401, s.nm_402
                    FROM normalized_spectra AS s
                    JOIN spectra_metadata AS m USING (source_id, spectrum_id, sample_name)
                    LIMIT 1
                    """
                ).fetchone()
                failure_paths = {
                    item[0]
                    for item in connection.execute(
                        "SELECT input_path FROM normalization_failures WHERE source_id = 'cabo_leaf_v2'"
                    ).fetchall()
                }
            finally:
                connection.close()

            self.assertEqual(row[0], "leaf_a")
            self.assertEqual(row[1], "csv_row_wide")
            self.assertAlmostEqual(row[2], 0.1)
            self.assertAlmostEqual(row[3], 0.2)
            self.assertAlmostEqual(row[4], 0.3)
            self.assertFalse(any("abs_spec.csv" in path for path in failure_paths))
            self.assertFalse(any("trans_spec.csv" in path for path in failure_paths))
            self.assertFalse(any("metadata_fields.csv" in path for path in failure_paths))

    def test_normalize_sources_filters_cabo_non_spectral_chemistry_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(manifest_path, [manifest_row(source_id="cabo_leaf_v2", name="CABO Leaf v2")])
            write_fetch_result(results_root / "cabo_leaf_v2" / "fetch-result.json", "cabo_leaf_v2")

            data_dir = results_root / "cabo_leaf_v2" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
            (data_dir / "ref_spec.csv").write_text(
                "sample_id,B208.9_mass,B249.8_mass,400,401,2400\nleaf_a,1.1,2.2,0.1,0.2,0.3\n",
                encoding="utf-8",
            )

            summary = normalize_sources(manifest_path, results_root, output_root)
            self.assertEqual(summary["normalized_spectra"], 1)

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                metadata_row = connection.execute(
                    "SELECT native_min_nm, native_max_nm, native_wavelength_count FROM spectra_metadata LIMIT 1"
                ).fetchone()
            finally:
                connection.close()

            self.assertEqual(metadata_row[0], 400.0)
            self.assertEqual(metadata_row[1], 2400.0)
            self.assertEqual(metadata_row[2], 3)

    def test_normalize_sources_skips_branch_tree_auxiliary_csv(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(
                manifest_path,
                [manifest_row(source_id="branch_tree_spectra_boreal_temperate", name="Branch Tree Spectra")],
            )
            write_fetch_result(
                results_root / "branch_tree_spectra_boreal_temperate" / "fetch-result.json",
                "branch_tree_spectra_boreal_temperate",
            )

            data_dir = results_root / "branch_tree_spectra_boreal_temperate" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
            (data_dir / "contact_probe_spectral_measurements.csv").write_text(
                "sample_id,400,401,2400\nbranch_a,0.1,0.2,0.3\n",
                encoding="utf-8",
            )
            (data_dir / "sampled_tree_descriptions.csv").write_text(
                "tree_height_m,branch_length_m\n10,1.2\n",
                encoding="utf-8",
            )

            summary = normalize_sources(manifest_path, results_root, output_root)
            self.assertEqual(summary["normalized_spectra"], 1)

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                sample_name = connection.execute("SELECT sample_name FROM spectra_metadata LIMIT 1").fetchone()[0]
                failure_paths = {
                    item[0]
                    for item in connection.execute(
                        "SELECT input_path FROM normalization_failures WHERE source_id = 'branch_tree_spectra_boreal_temperate'"
                    ).fetchall()
                }
            finally:
                connection.close()

            self.assertEqual(sample_name, "branch_a")
            self.assertFalse(any("sampled_tree_descriptions.csv" in path for path in failure_paths))

    def test_normalize_sources_skips_understory_icos_auxiliary_csv(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "manifests" / "sources.csv"
            results_root = root / "results"
            output_root = root / "normalized"

            write_manifest(
                manifest_path,
                [manifest_row(source_id="understory_icos_europe", name="Understory ICOS Europe")],
            )
            write_fetch_result(
                results_root / "understory_icos_europe" / "fetch-result.json",
                "understory_icos_europe",
            )

            data_dir = results_root / "understory_icos_europe" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
            (data_dir / "Pisek_et_al_2021_BG_supp_data.xlsx").write_text("", encoding="utf-8")
            (data_dir / "Pisek_et_al_2021_BG_Table1.csv").write_text(
                "site,latitude\nTEST,61.0\n",
                encoding="utf-8",
            )

            summary = normalize_sources(manifest_path, results_root, output_root)
            self.assertEqual(summary["normalized_spectra"], 0)

            connection = duckdb.connect(str(output_root / "db" / "normalized_catalog.duckdb"))
            try:
                failure_paths = {
                    item[0]
                    for item in connection.execute(
                        "SELECT input_path FROM normalization_failures WHERE source_id = 'understory_icos_europe'"
                    ).fetchall()
                }
            finally:
                connection.close()

            self.assertFalse(any("Pisek_et_al_2021_BG_Table1.csv" in path for path in failure_paths))


if __name__ == "__main__":
    unittest.main()
