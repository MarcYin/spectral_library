from __future__ import annotations

import csv
import io
import itertools
import json
import math
import re
import statistics
import sys
import zipfile
from array import array
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from .manifest import filter_sources, load_manifest


TARGET_START_NM = 400
TARGET_END_NM = 2500
TARGET_STEP_NM = 1
TARGET_WAVELENGTHS = tuple(range(TARGET_START_NM, TARGET_END_NM + 1, TARGET_STEP_NM))
SPECTRAL_COLUMN_PREFIX = "nm_"
SPECTRAL_COLUMNS = [f"{SPECTRAL_COLUMN_PREFIX}{wavelength}" for wavelength in TARGET_WAVELENGTHS]

LONG_WAVELENGTH_COLUMNS = ("wavelength", "wavelength_nm", "lambda", "wl")
LONG_VALUE_COLUMNS = ("r", "reflectance", "value", "y", "response")
PREFERRED_SAMPLE_COLUMNS = (
    "sample_name",
    "name",
    "spectra",
    "spectrum",
    "observation_id",
    "id_prog_spectrum",
    "id_spectrum_original",
    "sample_id",
    "id_unique",
    "index",
    "uniqueid",
    "id",
    "polygon_unit_id",
)

WAVELENGTH_HEADER_PATTERNS = (
    re.compile(r"^(?:x|wl|nm_)?(\d+(?:\.\d+)?)$", re.IGNORECASE),
    re.compile(r"^(?:spc\.?)(\d+(?:\.\d+)?)$", re.IGNORECASE),
    re.compile(r"^(\d+(?:\.\d+)?)\s*(?:nm|nanometers?)?$", re.IGNORECASE),
)
WAVELENGTH_TOKEN_PATTERN = re.compile(r"(?<!\d)(\d{3,4}(?:\.\d+)?)(?!\d)")


@dataclass(frozen=True)
class SpectrumRecord:
    source_id: str
    source_name: str
    ingest_role: str
    input_path: str
    parser: str
    sample_name: str
    wavelengths_nm: list[float]
    values: list[float]
    metadata: dict[str, str]
    value_scale_hint: float | None = None


def _normalize_header_name(value: str) -> str:
    return value.strip().strip('"').strip("'").lower()


def _parse_float(value: object | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    text = str(value).strip().strip('"').strip("'")
    if not text or text.lower() in {"na", "nan", "n/a", "null", "none"}:
        return None
    if text.endswith("%"):
        text = text[:-1].strip()
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _is_wavelength_header(label: str) -> bool:
    normalized = _normalize_header_name(label)
    return normalized in LONG_WAVELENGTH_COLUMNS or normalized.startswith("wavelength")


def _parse_wavelength_label(label: str) -> float | None:
    text = label.strip().strip('"').strip("'")
    for pattern in WAVELENGTH_HEADER_PATTERNS:
        match = pattern.match(text)
        if match:
            candidate = float(match.group(1))
            if 100 <= candidate <= 3000:
                return candidate
    for match in reversed(list(WAVELENGTH_TOKEN_PATTERN.finditer(text))):
        candidate = float(match.group(1))
        if 100 <= candidate <= 3000:
            return candidate
    return None


def _detect_value_scale(values: Iterable[float], hint: float | None = None) -> float:
    if hint and hint > 0:
        return hint

    finite = [abs(value) for value in values if math.isfinite(value)]
    if not finite:
        return 1.0

    max_value = max(finite)
    if max_value <= 1.5:
        return 1.0
    if max_value <= 100.5:
        return 100.0
    if max_value <= 1000.5:
        return 1000.0
    if max_value <= 10000.5:
        return 10000.0
    return 1.0


def _convert_wavelengths_to_nm(wavelengths: Iterable[float], units: str | None) -> list[float]:
    normalized_units = (units or "").strip().lower()
    wavelengths_list = list(wavelengths)
    if not wavelengths_list:
        return []

    multiplier = 1.0
    if "micro" in normalized_units or "microm" in normalized_units or normalized_units in {"um", "µm"}:
        multiplier = 1000.0
    elif "nano" in normalized_units or normalized_units == "nm":
        multiplier = 1.0
    elif normalized_units in {"m", "meter", "meters"} or (
        "meter" in normalized_units and "micro" not in normalized_units and "nano" not in normalized_units
    ):
        multiplier = 1_000_000_000.0
    elif max(wavelengths_list) <= 1e-5:
        multiplier = 1_000_000_000.0
    elif max(wavelengths_list) <= 10:
        multiplier = 1000.0

    return [value * multiplier for value in wavelengths_list]


def _median_spacing_nm(wavelengths_nm: list[float]) -> float | None:
    diffs = [
        wavelengths_nm[index + 1] - wavelengths_nm[index]
        for index in range(len(wavelengths_nm) - 1)
        if wavelengths_nm[index + 1] > wavelengths_nm[index]
    ]
    if not diffs:
        return None
    return statistics.median(diffs)


def _clean_spectrum(record: SpectrumRecord) -> tuple[list[float], list[float], float]:
    pairs = [
        (wavelength, value)
        for wavelength, value in zip(record.wavelengths_nm, record.values)
        if wavelength is not None and value is not None and math.isfinite(wavelength) and math.isfinite(value)
    ]
    if len(pairs) < 2:
        raise ValueError("spectrum has fewer than two finite wavelength/value pairs")

    pairs.sort(key=lambda pair: pair[0])
    deduped_wavelengths: list[float] = []
    deduped_values: list[float] = []
    for wavelength, value in pairs:
        if deduped_wavelengths and math.isclose(wavelength, deduped_wavelengths[-1], rel_tol=0.0, abs_tol=1e-9):
            deduped_values[-1] = value
            continue
        deduped_wavelengths.append(wavelength)
        deduped_values.append(value)

    if len(deduped_wavelengths) < 2:
        raise ValueError("spectrum collapsed to fewer than two unique wavelengths")

    scale = _detect_value_scale(deduped_values, record.value_scale_hint)
    if scale != 1.0:
        deduped_values = [value / scale for value in deduped_values]

    return deduped_wavelengths, deduped_values, scale


def _interpolate_to_grid(wavelengths_nm: list[float], values: list[float]) -> list[float | None]:
    normalized: list[float | None] = []
    index = 0
    final_index = len(wavelengths_nm) - 1

    for target in TARGET_WAVELENGTHS:
        if target < wavelengths_nm[0] or target > wavelengths_nm[-1]:
            normalized.append(None)
            continue

        while index + 1 < final_index and wavelengths_nm[index + 1] < target:
            index += 1

        left_wave = wavelengths_nm[index]
        left_value = values[index]

        if math.isclose(left_wave, target, rel_tol=0.0, abs_tol=1e-9):
            normalized.append(left_value)
            continue

        right_wave = wavelengths_nm[index + 1]
        right_value = values[index + 1]

        if math.isclose(right_wave, target, rel_tol=0.0, abs_tol=1e-9):
            normalized.append(right_value)
            continue

        fraction = (target - left_wave) / (right_wave - left_wave)
        normalized.append(left_value + (right_value - left_value) * fraction)

    return normalized


def _parse_envi_scalar(value: str) -> str | list[str]:
    text = value.strip()
    if text.startswith("{") and text.endswith("}"):
        inner = text[1:-1].strip()
        if not inner:
            return []
        return [item.strip() for item in inner.split(",")]
    return text


def _parse_envi_header(path: Path) -> dict[str, str | list[str]]:
    content = path.read_text(encoding="utf-8", errors="replace").splitlines()
    header: dict[str, str | list[str]] = {}
    current_key: str | None = None
    current_chunks: list[str] = []
    brace_depth = 0

    for raw_line in content:
        line = raw_line.strip()
        if not line or line.upper() == "ENVI":
            continue

        if current_key is None:
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            current_key = key.strip().lower()
            current_chunks = [value.strip()]
            brace_depth = value.count("{") - value.count("}")
            if brace_depth <= 0:
                header[current_key] = _parse_envi_scalar(" ".join(current_chunks))
                current_key = None
                current_chunks = []
                brace_depth = 0
            continue

        current_chunks.append(line)
        brace_depth += line.count("{") - line.count("}")
        if brace_depth <= 0:
            header[current_key] = _parse_envi_scalar(" ".join(current_chunks))
            current_key = None
            current_chunks = []
            brace_depth = 0

    if current_key is not None:
        header[current_key] = _parse_envi_scalar(" ".join(current_chunks))
    return header


def _extract_envi_float_list(header: dict[str, str | list[str]], key: str) -> list[float]:
    raw_value = header.get(key, [])
    items = raw_value if isinstance(raw_value, list) else [raw_value]
    values: list[float] = []
    for item in items:
        value = _parse_float(str(item))
        if value is not None:
            values.append(value)
    return values


def _extract_envi_string_list(header: dict[str, str | list[str]], key: str) -> list[str]:
    raw_value = header.get(key, [])
    items = raw_value if isinstance(raw_value, list) else [raw_value]
    return [str(item).strip() for item in items if str(item).strip()]


def _resolve_envi_data_path(header_path: Path, header: dict[str, str | list[str]]) -> Path:
    stem_without_suffix = header_path.with_suffix("")
    if stem_without_suffix.exists():
        return stem_without_suffix

    spectral_library_path = header_path.with_suffix(".sli")
    if spectral_library_path.exists():
        return spectral_library_path

    img_path = header_path.with_suffix(".img")
    if img_path.exists():
        return img_path

    data_file = str(header.get("data file", "")).strip()
    if data_file:
        explicit_path = header_path.parent / data_file
        if explicit_path.exists():
            return explicit_path

    raise FileNotFoundError(f"could not resolve ENVI binary companion for {header_path.name}")


def _envi_shape(header: dict[str, str | list[str]]) -> tuple[int, int]:
    samples = int(float(str(header.get("samples", "0"))))
    lines = int(float(str(header.get("lines", "0"))))
    bands = int(float(str(header.get("bands", "0"))))
    file_type = str(header.get("file type", "")).lower()
    interleave = str(header.get("interleave", "")).lower()

    if "spectral library" in file_type:
        return lines, samples
    if interleave == "bip" and samples == 1:
        return lines, bands
    if interleave == "bsq" and bands == 1:
        return lines, samples
    raise ValueError(f"unsupported ENVI layout: file_type={file_type or 'unknown'} interleave={interleave or 'unknown'}")


def _envi_typecode(data_type: int) -> str:
    mapping = {
        1: "B",
        2: "h",
        3: "i",
        4: "f",
        5: "d",
        12: "H",
        13: "I",
        14: "q",
        15: "Q",
    }
    if data_type not in mapping:
        raise ValueError(f"unsupported ENVI data type: {data_type}")
    return mapping[data_type]


def _stringify_cell(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _input_stem(input_path: str | Path) -> str:
    return Path(str(input_path).split("::")[-1]).stem


def _pick_sample_name(row: dict[str, object | None], row_index: int) -> str:
    normalized_lookup = {_normalize_header_name(key): key for key in row}
    for candidate in PREFERRED_SAMPLE_COLUMNS:
        key = normalized_lookup.get(candidate)
        if not key:
            continue
        value = _stringify_cell(row.get(key))
        if value and value.lower() not in {"na", "nan", "null", "none"}:
            return value
    for key, value in row.items():
        normalized_key = _normalize_header_name(key)
        if _parse_wavelength_label(key) is not None or _is_wavelength_header(normalized_key):
            continue
        if normalized_key in LONG_VALUE_COLUMNS:
            continue
        text = _stringify_cell(value)
        if text and text.lower() not in {"na", "nan", "null", "none"}:
            return text
    return f"spectrum_{row_index:05d}"


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        if sample.count(";") > sample.count(","):
            return ";"
        if sample.count("\t") > sample.count(","):
            return "\t"
        return ","


def _strip_tabular_comments(lines: Sequence[str]) -> list[str]:
    cleaned: list[str] = []
    in_block_comment = False
    for raw_line in lines:
        line = raw_line.rstrip("\n")
        stripped = line.strip()
        if in_block_comment:
            if "*/" in stripped:
                in_block_comment = False
            continue
        if stripped.startswith("/*"):
            if "*/" not in stripped:
                in_block_comment = True
            continue
        if stripped.startswith("#"):
            continue
        cleaned.append(line)
    return cleaned


def _find_tabular_header(lines: list[str], delimiter: str) -> tuple[int, list[str]]:
    for index, line in enumerate(lines):
        if not line.strip():
            continue
        parsed = next(csv.reader([line], delimiter=delimiter))
        wavelength_hits = sum(1 for value in parsed if _parse_wavelength_label(value) is not None)
        if parsed and _is_wavelength_header(_stringify_cell(parsed[0])):
            return index, parsed
        if wavelength_hits >= 3:
            return index, parsed
    first = next(csv.reader([lines[0]], delimiter=delimiter)) if lines else []
    return 0, first


def _iter_row_wide_spectra(
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
    header: list[str],
    reader: Iterable[dict[str, object | None]],
    *,
    parser_name: str = "csv_row_wide",
) -> Iterator[SpectrumRecord]:
    wavelength_fields = [(field, _parse_wavelength_label(field)) for field in header]
    spectral_fields = [(field, wavelength) for field, wavelength in wavelength_fields if wavelength is not None]
    spectral_names = {field for field, _ in spectral_fields}
    row_index = 0

    for row in reader:
        row_index += 1
        wavelengths_nm = [wavelength for _, wavelength in spectral_fields]
        values = [_parse_float(row.get(field)) for field, _ in spectral_fields]
        if not any(value is not None for value in values):
            continue

        metadata = {
            key: text
            for key, value in row.items()
            if key not in spectral_names and (text := _stringify_cell(value))
        }
        sample_name = _pick_sample_name(row, row_index)
        yield SpectrumRecord(
            source_id=source_id,
            source_name=source_name,
            ingest_role=ingest_role,
            input_path=str(input_path),
            parser=parser_name,
            sample_name=sample_name,
            wavelengths_nm=wavelengths_nm,
            values=[value if value is not None else math.nan for value in values],
            metadata=metadata,
        )


def _iter_column_wise_spectra(
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
    header: list[str],
    rows: Iterable[Sequence[object | None]],
    *,
    parser_name: str = "csv_column_wide",
) -> Iterator[SpectrumRecord]:
    sample_names = [_stringify_cell(value) for value in header[1:]]
    if not sample_names:
        return

    wavelengths_nm: list[float] = []
    sample_values = [[] for _ in sample_names]

    for row in rows:
        if not row or not _stringify_cell(row[0]):
            continue
        wavelength = _parse_float(row[0])
        if wavelength is None:
            continue
        wavelengths_nm.append(wavelength)
        padded = list(row[1:]) + [None] * max(0, len(sample_names) - len(row[1:]))
        for index, value in enumerate(padded[: len(sample_names)]):
            sample_values[index].append(_parse_float(value))

    for index, (sample_name, values) in enumerate(zip(sample_names, sample_values), start=1):
        if not any(value is not None for value in values):
            continue
        yield SpectrumRecord(
            source_id=source_id,
            source_name=source_name,
            ingest_role=ingest_role,
            input_path=str(input_path),
            parser=parser_name,
            sample_name=sample_name or f"unnamed_sample_{index:05d}",
            wavelengths_nm=wavelengths_nm,
            values=[value if value is not None else math.nan for value in values],
            metadata={},
        )


def _iter_long_table_spectra(
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
    header: list[str],
    rows: Iterable[dict[str, object | None]],
    *,
    parser_name: str = "csv_long_table",
) -> Iterator[SpectrumRecord]:
    normalized_lookup = {_normalize_header_name(value): value for value in header}
    wavelength_key = next(
        (value for name, value in normalized_lookup.items() if _is_wavelength_header(name)),
        None,
    )
    value_key = next(
        (normalized_lookup[name] for name in LONG_VALUE_COLUMNS if name in normalized_lookup),
        None,
    )
    if not wavelength_key or not value_key:
        return

    sample_key = next(
        (normalized_lookup[name] for name in PREFERRED_SAMPLE_COLUMNS if name in normalized_lookup and name != "wavelength"),
        None,
    )

    current_sample = ""
    current_wavelengths: list[float] = []
    current_values: list[float] = []
    current_metadata: dict[str, str] = {}
    sample_counter = 0

    def flush() -> SpectrumRecord | None:
        nonlocal sample_counter
        if not current_sample or not current_wavelengths:
            return None
        sample_counter += 1
        return SpectrumRecord(
            source_id=source_id,
            source_name=source_name,
            ingest_role=ingest_role,
            input_path=str(input_path),
            parser=parser_name,
            sample_name=current_sample,
            wavelengths_nm=list(current_wavelengths),
            values=list(current_values),
            metadata=dict(current_metadata),
        )

    row_index = 0
    for row in rows:
        row_index += 1
        wavelength = _parse_float(row.get(wavelength_key))
        value = _parse_float(row.get(value_key))
        if wavelength is None or value is None:
            continue

        sample_name = _stringify_cell(row.get(sample_key)) if sample_key else f"spectrum_{row_index:05d}"
        if not sample_name:
            sample_name = f"spectrum_{row_index:05d}"

        if current_sample and sample_name != current_sample:
            record = flush()
            if record is not None:
                yield record
            current_wavelengths.clear()
            current_values.clear()
            current_metadata.clear()

        if sample_name != current_sample:
            current_sample = sample_name
            current_metadata = {
                key: text
                for key, value in row.items()
                if key not in {wavelength_key, value_key} and (text := _stringify_cell(value))
            }

        current_wavelengths.append(wavelength)
        current_values.append(value)

    record = flush()
    if record is not None:
        yield record


def _iter_tabular_spectra_from_lines(
    input_path: str | Path,
    raw_lines: Sequence[str],
    source_id: str,
    source_name: str,
    ingest_role: str,
    *,
    parser_prefix: str = "csv",
) -> Iterator[SpectrumRecord]:
    lines = _strip_tabular_comments(raw_lines)
    if not any(line.strip() for line in lines):
        return

    preview = "\n".join(lines[:80])
    delimiter = _sniff_delimiter(preview)
    preview_lines = lines[:80] or [""]
    header_index, _ = _find_tabular_header(preview_lines, delimiter)
    data_lines = lines[header_index:]
    if not data_lines:
        return

    reader = csv.reader(data_lines, delimiter=delimiter)
    header = next(reader, [])
    if not header:
        return

    normalized = [_normalize_header_name(_stringify_cell(value)) for value in header]
    wavelength_hits = sum(1 for value in header if _parse_wavelength_label(_stringify_cell(value)) is not None)

    if (
        header_index > 0
        and normalized
        and _is_wavelength_header(_stringify_cell(header[0]))
        and wavelength_hits >= 3
    ):
        previous_header = next(csv.reader([lines[header_index - 1]], delimiter=delimiter))
        sample_name_header = _stringify_cell(previous_header[0] if previous_header else None)
        if sample_name_header and not _is_wavelength_header(sample_name_header):
            combined_header = [sample_name_header, *[_stringify_cell(value) for value in header[1:]]]
            dict_reader = csv.DictReader(data_lines[1:], fieldnames=combined_header, delimiter=delimiter)
            yield from _iter_row_wide_spectra(
                input_path,
                source_id,
                source_name,
                ingest_role,
                combined_header,
                dict_reader,
                parser_name=f"{parser_prefix}_row_wide",
            )
            return

    if wavelength_hits < 3 and len(data_lines) >= 2:
        second_header = next(csv.reader([data_lines[1]], delimiter=delimiter))
        second_hits = sum(1 for value in second_header if _parse_wavelength_label(_stringify_cell(value)) is not None)
        if second_header and _is_wavelength_header(_stringify_cell(second_header[0])) and second_hits >= 3:
            sample_name_header = _stringify_cell(header[0]) or "sample_name"
            combined_header = [sample_name_header, *[_stringify_cell(value) for value in second_header[1:]]]
            dict_reader = csv.DictReader(data_lines[2:], fieldnames=combined_header, delimiter=delimiter)
            yield from _iter_row_wide_spectra(
                input_path,
                source_id,
                source_name,
                ingest_role,
                combined_header,
                dict_reader,
                parser_name=f"{parser_prefix}_row_wide",
            )
            return

    if normalized and _is_wavelength_header(_stringify_cell(header[0])) and wavelength_hits <= 1:
        yield from _iter_column_wise_spectra(
            input_path,
            source_id,
            source_name,
            ingest_role,
            [_stringify_cell(value) for value in header],
            reader,
            parser_name=f"{parser_prefix}_column_wide",
        )
        return

    dict_reader = csv.DictReader(data_lines[1:], fieldnames=[_stringify_cell(value) for value in header], delimiter=delimiter)
    if any(_is_wavelength_header(value) for value in normalized) and any(value in LONG_VALUE_COLUMNS for value in normalized):
        yield from _iter_long_table_spectra(
            input_path,
            source_id,
            source_name,
            ingest_role,
            [_stringify_cell(value) for value in header],
            dict_reader,
            parser_name=f"{parser_prefix}_long_table",
        )
        return

    if wavelength_hits >= 3:
        yield from _iter_row_wide_spectra(
            input_path,
            source_id,
            source_name,
            ingest_role,
            [_stringify_cell(value) for value in header],
            dict_reader,
            parser_name=f"{parser_prefix}_row_wide",
        )


def _iter_tabular_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        lines = handle.read().splitlines()
    yield from _iter_tabular_spectra_from_lines(path, lines, source_id, source_name, ingest_role)


def _iter_ecostress_text_spectra_from_lines(
    lines: Sequence[str],
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    metadata: dict[str, str] = {}
    wavelengths: list[float] = []
    values: list[float] = []
    in_data_section = False

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            if metadata:
                in_data_section = True
            continue
        if not in_data_section and ":" in line:
            key, value = line.split(":", 1)
            metadata[key.strip()] = value.strip()
            continue
        if not in_data_section:
            continue
        parts = [part for part in re.split(r"[\s,;\t]+", line) if part]
        if len(parts) < 2:
            continue
        wavelength = _parse_float(parts[0])
        value = _parse_float(parts[1])
        if wavelength is None or value is None:
            continue
        wavelengths.append(wavelength)
        values.append(value)

    if not metadata or not wavelengths:
        return

    x_units = metadata.get("X Units", "")
    y_units = metadata.get("Y Units", "")
    scale_hint = 100.0 if "percent" in y_units.lower() else None

    yield SpectrumRecord(
        source_id=source_id,
        source_name=source_name,
        ingest_role=ingest_role,
        input_path=str(input_path),
        parser="ecostress_txt",
        sample_name=metadata.get("Name", _input_stem(input_path)),
        wavelengths_nm=_convert_wavelengths_to_nm(wavelengths, x_units),
        values=values,
        metadata=metadata,
        value_scale_hint=scale_hint,
    )


def _iter_ecostress_text_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        lines = handle.read().splitlines()
    yield from _iter_ecostress_text_spectra_from_lines(lines, path, source_id, source_name, ingest_role)


def _iter_usgs_text_spectra_from_lines(
    lines: Sequence[str],
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    if not lines:
        return
    title = lines[0].strip()
    if "Record=" not in title:
        return

    raw_values = [_parse_float(line) for line in lines[1:] if line.strip()]
    if len(raw_values) < 1000:
        return

    wavelengths_nm = list(range(350, 350 + len(raw_values)))
    values = [math.nan if value is None or abs(value) > 1e20 else value for value in raw_values]
    sample_name = title.split(":", 1)[-1].strip() or _input_stem(input_path)

    yield SpectrumRecord(
        source_id=source_id,
        source_name=source_name,
        ingest_role=ingest_role,
        input_path=str(input_path),
        parser="usgs_ascii",
        sample_name=sample_name,
        wavelengths_nm=wavelengths_nm,
        values=values,
        metadata={"title": title},
    )


def _iter_pair_text_spectra_from_lines(
    lines: Sequence[str],
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    non_empty = [line.strip() for line in lines if line.strip()]
    if len(non_empty) < 3:
        return

    header_index = next(
        (
            index
            for index, line in enumerate(non_empty[:5])
            if "wavelength" in line.lower() and any(token in line.lower() for token in ("reflectance", "response", "value", "rrs", "refl"))
        ),
        None,
    )
    if header_index is None:
        return

    sample_name = non_empty[0] if header_index > 0 else _input_stem(input_path)
    header_line = non_empty[header_index]
    wavelengths: list[float] = []
    values: list[float] = []

    for line in non_empty[header_index + 1 :]:
        parts = [part for part in re.split(r"[\s,;\t]+", line) if part]
        if len(parts) < 2:
            continue
        wavelength = _parse_float(parts[0])
        value = _parse_float(parts[1])
        if wavelength is None or value is None:
            continue
        wavelengths.append(wavelength)
        values.append(value)

    if not wavelengths:
        return

    yield SpectrumRecord(
        source_id=source_id,
        source_name=source_name,
        ingest_role=ingest_role,
        input_path=str(input_path),
        parser="text_pairs",
        sample_name=sample_name,
        wavelengths_nm=_convert_wavelengths_to_nm(wavelengths, "nm" if "nm" in header_line.lower() else ""),
        values=values,
        metadata={"header": header_line},
    )


def _iter_textual_spectra_from_lines(
    lines: Sequence[str],
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    factories = (
        lambda: _iter_usgs_text_spectra_from_lines(lines, input_path, source_id, source_name, ingest_role),
        lambda: _iter_ecostress_text_spectra_from_lines(lines, input_path, source_id, source_name, ingest_role),
        lambda: _iter_pair_text_spectra_from_lines(lines, input_path, source_id, source_name, ingest_role),
        lambda: _iter_tabular_spectra_from_lines(input_path, lines, source_id, source_name, ingest_role),
    )
    for factory in factories:
        iterator = iter(factory())
        first = next(iterator, None)
        if first is None:
            continue
        yield first
        yield from iterator
        return


def _decode_text(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def _netcdf_array_to_list(values: object) -> list[float]:
    if hasattr(values, "filled"):
        values = values.filled(math.nan)
    if hasattr(values, "tolist"):
        values = values.tolist()
    return [math.nan if (parsed := _parse_float(value)) is None else parsed for value in values]


def _iter_netcdf_spectra_bytes(
    input_path: str | Path,
    payload: bytes,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    from netCDF4 import Dataset

    dataset = Dataset("inmemory.nc", memory=payload)
    try:
        wavelength_name = next(
            (name for name in dataset.variables if _is_wavelength_header(_normalize_header_name(name))),
            None,
        )
        if not wavelength_name:
            raise ValueError("netCDF file is missing a wavelength variable")
        wavelength_var = dataset.variables[wavelength_name]
        wavelengths_nm = _convert_wavelengths_to_nm(_netcdf_array_to_list(wavelength_var[:]), getattr(wavelength_var, "units", ""))

        preferred_names = ("reflectance", "rrs", "rw", "refl")
        value_name = next(
            (
                name
                for name in dataset.variables
                if _normalize_header_name(name) in preferred_names and "wavelength" in dataset.variables[name].dimensions
            ),
            None,
        )
        if value_name is None:
            value_name = next(
                (
                    name
                    for name in dataset.variables
                    if "reflectance" in _normalize_header_name(name)
                    and "std" not in _normalize_header_name(name)
                    and "wavelength" in dataset.variables[name].dimensions
                ),
                None,
            )
        if not value_name:
            raise ValueError("netCDF file is missing a reflectance variable")

        value_var = dataset.variables[value_name]
        metadata = {
            "value_variable": value_name,
            "wavelength_variable": wavelength_name,
            "netcdf_dimensions": json.dumps({name: len(dimension) for name, dimension in dataset.dimensions.items()}, sort_keys=True),
        }
        for attr_name in dataset.ncattrs():
            value = getattr(dataset, attr_name)
            metadata[f"attr:{attr_name}"] = _stringify_cell(value)

        if len(value_var.dimensions) == 1:
            yield SpectrumRecord(
                source_id=source_id,
                source_name=source_name,
                ingest_role=ingest_role,
                input_path=str(input_path),
                parser="netcdf_reflectance",
                sample_name=_input_stem(input_path),
                wavelengths_nm=wavelengths_nm,
                values=_netcdf_array_to_list(value_var[:]),
                metadata=metadata,
            )
            return

        if len(value_var.dimensions) == 2 and value_var.dimensions[-1] == wavelength_name:
            observation_count = value_var.shape[0]
            for index in range(observation_count):
                sample_metadata = dict(metadata)
                sample_metadata["obs_index"] = str(index)
                yield SpectrumRecord(
                    source_id=source_id,
                    source_name=source_name,
                    ingest_role=ingest_role,
                    input_path=str(input_path),
                    parser="netcdf_reflectance",
                    sample_name=f"{_input_stem(input_path)}_{index + 1:05d}",
                    wavelengths_nm=wavelengths_nm,
                    values=_netcdf_array_to_list(value_var[index, :]),
                    metadata=sample_metadata,
                )
            return

        raise ValueError(f"unsupported netCDF reflectance dimensions {value_var.dimensions!r}")
    finally:
        dataset.close()


def _iter_netcdf_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    yield from _iter_netcdf_spectra_bytes(path, path.read_bytes(), source_id, source_name, ingest_role)


def _iter_zip_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    supported_suffixes = {".csv", ".txt", ".tab", ".nc"}
    preferred_tokens = ("spectra", "spectrum", "reflectance", "refl", "rrs")
    ancillary_tokens = ("metadata", "quality", "readme", "license")
    with zipfile.ZipFile(path) as archive:
        member_names = sorted(
            archive.namelist(),
            key=lambda member_name: (
                0 if any(token in member_name.lower() for token in preferred_tokens) else 1,
                member_name.lower(),
            ),
        )
        for member_name in member_names:
            member_path = Path(member_name)
            if member_name.endswith("/") or member_path.name.startswith(".") or member_path.parts[:1] == ("__MACOSX",):
                continue
            suffix = member_path.suffix.lower()
            if suffix not in supported_suffixes:
                continue
            member_name_normalized = member_name.lower()
            if any(token in member_name_normalized for token in ancillary_tokens) and not any(
                token in member_name_normalized for token in preferred_tokens
            ):
                continue
            input_path = f"{path}::{member_name}"
            payload = archive.read(member_name)
            if suffix == ".nc":
                yield from _iter_netcdf_spectra_bytes(input_path, payload, source_id, source_name, ingest_role)
            else:
                lines = _decode_text(payload).splitlines()
                yield from _iter_textual_spectra_from_lines(lines, input_path, source_id, source_name, ingest_role)


def _iter_xlsx_band_matrix_sheet(
    path: Path,
    sheet_name: str,
    preview_rows: list[tuple[object | None, ...]],
    header_index: int,
    source_id: str,
    source_name: str,
    ingest_role: str,
) -> Iterator[SpectrumRecord]:
    import openpyxl

    header_row = preview_rows[header_index]
    band_column = next(
        index
        for index, value in enumerate(header_row)
        if _normalize_header_name(_stringify_cell(value)) == "band"
    )
    sample_columns = [
        index
        for index in range(band_column + 1, len(header_row))
        if any(_stringify_cell(row[index] if index < len(row) else None) for row in preview_rows[header_index + 1 : header_index + 6])
    ]
    if not sample_columns:
        return

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        sample_metadata_by_column: dict[int, dict[str, str]] = {}
        for column_index in sample_columns:
            metadata: dict[str, str] = {"sheet": sheet_name}
            for row in preview_rows[:header_index]:
                if not row:
                    continue
                label = _stringify_cell(row[0] if len(row) > 0 else None)
                value = _stringify_cell(row[column_index] if column_index < len(row) else None)
                if label and value:
                    metadata[_normalize_header_name(label)] = value
            sample_metadata_by_column[column_index] = metadata

        wavelengths_nm: list[float] = []
        sample_values = {column_index: [] for column_index in sample_columns}
        for row in worksheet.iter_rows(min_row=header_index + 2, values_only=True):
            wavelength = _parse_float(row[band_column] if band_column < len(row) else None)
            if wavelength is None:
                continue
            wavelengths_nm.append(wavelength)
            for column_index in sample_columns:
                sample_values[column_index].append(_parse_float(row[column_index] if column_index < len(row) else None))

        for column_index in sample_columns:
            metadata = sample_metadata_by_column[column_index]
            plot_id = metadata.get("plot id")
            julian_day = metadata.get("julian day")
            sample_name_parts = [plot_id or f"{sheet_name}_{column_index + 1:03d}"]
            if julian_day:
                sample_name_parts.append(f"jd{julian_day}")
            yield SpectrumRecord(
                source_id=source_id,
                source_name=source_name,
                ingest_role=ingest_role,
                input_path=f"{path}::{sheet_name}",
                parser="xlsx_band_matrix",
                sample_name="_".join(sample_name_parts),
                wavelengths_nm=wavelengths_nm,
                values=[value if value is not None else math.nan for value in sample_values[column_index]],
                metadata=metadata,
            )
    finally:
        workbook.close()


def _iter_xlsx_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    for sheet_name in sheet_names:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name]
            preview_rows = [tuple(row) for row in worksheet.iter_rows(min_row=1, max_row=80, values_only=True)]
        finally:
            workbook.close()

        band_matrix_index = next(
            (
                index
                for index, row in enumerate(preview_rows)
                if [_normalize_header_name(_stringify_cell(value)) for value in row].count("reflectance") >= 3
                and "band" in [_normalize_header_name(_stringify_cell(value)) for value in row]
            ),
            None,
        )
        if band_matrix_index is not None:
            yield from _iter_xlsx_band_matrix_sheet(
                path,
                sheet_name,
                preview_rows,
                band_matrix_index,
                source_id,
                source_name,
                ingest_role,
            )
            continue

        row_wide_index = next(
            (
                index
                for index, row in enumerate(preview_rows)
                if sum(1 for value in row if _parse_wavelength_label(_stringify_cell(value)) is not None) >= 3
            ),
            None,
        )
        if row_wide_index is not None:
            header = [_stringify_cell(value) for value in preview_rows[row_wide_index]]

            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                worksheet = workbook[sheet_name]

                def row_dicts() -> Iterator[dict[str, object | None]]:
                    for row in worksheet.iter_rows(min_row=row_wide_index + 2, values_only=True):
                        yield {
                            header[index]: row[index] if index < len(row) else None
                            for index in range(len(header))
                        }

                yield from _iter_row_wide_spectra(
                    f"{path}::{sheet_name}",
                    source_id,
                    source_name,
                    ingest_role,
                    header,
                    row_dicts(),
                    parser_name="xlsx_row_wide",
                )
            finally:
                workbook.close()
            continue


def _iter_dataframe_spectra(
    frame: object,
    input_path: str | Path,
    source_id: str,
    source_name: str,
    ingest_role: str,
    *,
    parser_name: str = "rds_dataframe",
) -> Iterator[SpectrumRecord]:
    header = [str(column) for column in frame.columns]

    def row_dicts() -> Iterator[dict[str, object | None]]:
        for row in frame.itertuples(index=False, name=None):
            yield {header[index]: row[index] for index in range(len(header))}

    yield from _iter_row_wide_spectra(
        input_path,
        source_id,
        source_name,
        ingest_role,
        header,
        row_dicts(),
        parser_name=parser_name,
    )


def _iter_rds_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    import rdata

    payload = rdata.read_rds(path, expand_altrep=False)
    if hasattr(payload, "columns") and hasattr(payload, "itertuples"):
        yield from _iter_dataframe_spectra(payload, path, source_id, source_name, ingest_role)
        return
    if isinstance(payload, dict):
        for key, value in payload.items():
            if hasattr(value, "columns") and hasattr(value, "itertuples"):
                yield from _iter_dataframe_spectra(
                    value,
                    f"{path}::{key}",
                    source_id,
                    source_name,
                    ingest_role,
                )
                return
    raise ValueError(f"unsupported RDS payload type {type(payload).__name__}")


def _iter_envi_spectra(path: Path, source_id: str, source_name: str, ingest_role: str) -> Iterator[SpectrumRecord]:
    header = _parse_envi_header(path)
    data_path = _resolve_envi_data_path(path, header)
    wavelengths = _extract_envi_float_list(header, "wavelength")
    if not wavelengths:
        raise ValueError("ENVI header is missing wavelength values")
    wavelengths_nm = _convert_wavelengths_to_nm(wavelengths, str(header.get("wavelength units", "")))

    data_type = int(float(str(header.get("data type", "0"))))
    typecode = _envi_typecode(data_type)
    spectrum_count, wavelength_count = _envi_shape(header)
    expected_value_count = spectrum_count * wavelength_count

    payload = data_path.read_bytes()
    values = array(typecode)
    values.frombytes(payload)
    if str(header.get("byte order", "0")).strip() == "1" and sys.byteorder == "little":
        values.byteswap()
    if len(values) != expected_value_count:
        raise ValueError(
            f"ENVI binary length mismatch for {data_path.name}: expected {expected_value_count} values, found {len(values)}"
        )

    scale_hint = _parse_float(str(header.get("reflectance scale factor", "")))
    spectra_names = _extract_envi_string_list(header, "spectra names")
    file_type = str(header.get("file type", ""))
    interleave = str(header.get("interleave", ""))

    for index in range(spectrum_count):
        start = index * wavelength_count
        stop = start + wavelength_count
        sample_name = spectra_names[index] if index < len(spectra_names) else f"spectrum_{index + 1:05d}"
        yield SpectrumRecord(
            source_id=source_id,
            source_name=source_name,
            ingest_role=ingest_role,
            input_path=str(path),
            parser="envi_binary",
            sample_name=sample_name,
            wavelengths_nm=wavelengths_nm,
            values=[float(value) for value in values[start:stop]],
            metadata={
                "header_path": str(path),
                "data_path": str(data_path),
                "file_type": file_type,
                "interleave": interleave,
            },
            value_scale_hint=scale_hint,
        )


def _open_csv_writer(path: Path, columns: list[str]) -> tuple[object, csv.writer]:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = path.open("w", encoding="utf-8", newline="")
    writer = csv.writer(handle)
    writer.writerow(columns)
    return handle, writer


def _write_failure(writer: csv.writer, source_id: str, input_path: str, parser: str, stage: str, reason: str) -> None:
    writer.writerow([source_id, input_path, parser, stage, reason])


def _write_grid(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["wavelength_nm", "column_name"])
        for wavelength in TARGET_WAVELENGTHS:
            writer.writerow([wavelength, f"{SPECTRAL_COLUMN_PREFIX}{wavelength}"])


def _file_sort_key(path: Path) -> tuple[int, str]:
    suffix = path.suffix.lower()
    priority = {
        ".hdr": 0,
        ".zip": 1,
        ".xlsx": 2,
        ".xls": 2,
        ".rds": 2,
        ".nc": 2,
        ".txt": 3,
        ".tab": 4,
        ".csv": 5,
        ".sli": 6,
    }.get(suffix, 7)
    return priority, str(path)


def _field_type_spec(field_names: list[str], double_fields: Iterable[str] = ()) -> str:
    numeric = set(double_fields)
    parts = []
    for field_name in field_names:
        field_type = "DOUBLE" if field_name in numeric else "VARCHAR"
        parts.append(f"'{field_name}': '{field_type}'")
    return "{" + ", ".join(parts) + "}"


def _build_database(output_root: Path) -> None:
    import duckdb

    tabular_dir = output_root / "tabular"
    parquet_dir = output_root / "parquet"
    db_dir = output_root / "db"
    parquet_dir.mkdir(parents=True, exist_ok=True)
    db_dir.mkdir(parents=True, exist_ok=True)

    metadata_columns = [
        "source_id",
        "source_name",
        "ingest_role",
        "spectrum_id",
        "sample_name",
        "input_path",
        "parser",
        "native_wavelength_count",
        "native_min_nm",
        "native_max_nm",
        "native_spacing_nm",
        "value_scale_applied",
        "normalized_points",
        "metadata_json",
    ]
    failure_columns = ["source_id", "input_path", "parser", "stage", "reason"]
    summary_columns = [
        "source_id",
        "source_name",
        "ingest_role",
        "normalized_spectra",
        "failure_count",
        "parsers",
    ]
    normalized_columns = ["source_id", "spectrum_id", "sample_name", *SPECTRAL_COLUMNS]

    connection = duckdb.connect(str(db_dir / "normalized_catalog.duckdb"))
    try:
        connection.execute(
            f"""
            CREATE OR REPLACE TABLE wavelength_grid AS
            SELECT * FROM read_csv(
                ?,
                header = true,
                columns = {_field_type_spec(['wavelength_nm', 'column_name'], ['wavelength_nm'])}
            )
            """,
            [str(tabular_dir / "wavelength_grid.csv")],
        )
        connection.execute(
            f"""
            CREATE OR REPLACE TABLE spectra_metadata AS
            SELECT * FROM read_csv(
                ?,
                header = true,
                columns = {_field_type_spec(
                    metadata_columns,
                    ['native_wavelength_count', 'native_min_nm', 'native_max_nm', 'native_spacing_nm', 'value_scale_applied', 'normalized_points'],
                )},
                nullstr = ''
            )
            """,
            [str(tabular_dir / "spectra_metadata.csv")],
        )
        connection.execute(
            f"""
            CREATE OR REPLACE TABLE normalization_failures AS
            SELECT * FROM read_csv(
                ?,
                header = true,
                columns = {_field_type_spec(failure_columns)},
                nullstr = ''
            )
            """,
            [str(tabular_dir / "normalization_failures.csv")],
        )
        connection.execute(
            f"""
            CREATE OR REPLACE TABLE source_summary AS
            SELECT * FROM read_csv(
                ?,
                header = true,
                columns = {_field_type_spec(summary_columns, ['normalized_spectra', 'failure_count'])},
                nullstr = ''
            )
            """,
            [str(tabular_dir / "source_summary.csv")],
        )
        connection.execute(
            f"""
            CREATE OR REPLACE TABLE normalized_spectra AS
            SELECT * FROM read_csv(
                ?,
                header = true,
                columns = {_field_type_spec(normalized_columns, SPECTRAL_COLUMNS)},
                nullstr = ''
            )
            """,
            [str(tabular_dir / "normalized_spectra.csv")],
        )
        connection.execute(
            """
            CREATE OR REPLACE VIEW source_normalization_status AS
            SELECT
              ss.source_id,
              ss.source_name,
              ss.ingest_role,
              ss.normalized_spectra,
              ss.failure_count,
              ss.parsers
            FROM source_summary ss
            ORDER BY ss.source_id
            """
        )
        for table_name in (
            "wavelength_grid",
            "spectra_metadata",
            "normalized_spectra",
            "normalization_failures",
            "source_summary",
        ):
            connection.execute(
                f"COPY {table_name} TO ? (FORMAT PARQUET)",
                [str(parquet_dir / f"{table_name}.parquet")],
            )
    finally:
        connection.close()


def normalize_sources(
    manifest_path: Path,
    results_root: Path,
    output_root: Path,
    *,
    source_ids: list[str] | None = None,
    limit: int = 0,
) -> dict[str, object]:
    tabular_dir = output_root / "tabular"
    tabular_dir.mkdir(parents=True, exist_ok=True)

    metadata_handle, metadata_writer = _open_csv_writer(
        tabular_dir / "spectra_metadata.csv",
        [
            "source_id",
            "source_name",
            "ingest_role",
            "spectrum_id",
            "sample_name",
            "input_path",
            "parser",
            "native_wavelength_count",
            "native_min_nm",
            "native_max_nm",
            "native_spacing_nm",
            "value_scale_applied",
            "normalized_points",
            "metadata_json",
        ],
    )
    normalized_handle, normalized_writer = _open_csv_writer(
        tabular_dir / "normalized_spectra.csv",
        ["source_id", "spectrum_id", "sample_name", *SPECTRAL_COLUMNS],
    )
    failure_handle, failure_writer = _open_csv_writer(
        tabular_dir / "normalization_failures.csv",
        ["source_id", "input_path", "parser", "stage", "reason"],
    )
    summary_handle, summary_writer = _open_csv_writer(
        tabular_dir / "source_summary.csv",
        ["source_id", "source_name", "ingest_role", "normalized_spectra", "failure_count", "parsers"],
    )
    _write_grid(tabular_dir / "wavelength_grid.csv")

    manifest_records = load_manifest(manifest_path)
    selected = filter_sources(manifest_records, source_ids=source_ids)
    if limit > 0:
        selected = selected[:limit]

    parser_counts: Counter[str] = Counter()
    total_failures = 0
    total_spectra = 0
    normalized_sources = 0
    downloaded_sources = 0

    try:
        for source in selected:
            source_dir = results_root / source.source_id
            fetch_result_path = source_dir / "fetch-result.json"
            if not fetch_result_path.exists():
                continue

            fetch_result = json.loads(fetch_result_path.read_text(encoding="utf-8"))
            if fetch_result.get("status") != "downloaded":
                continue

            downloaded_sources += 1
            data_dir = source_dir / "data"
            source_spectra = 0
            source_failures = 0
            source_parsers: Counter[str] = Counter()
            spectrum_index = 0
            handled_paths: set[Path] = set()

            candidate_paths: list[Path] = []
            if data_dir.exists():
                candidate_paths.extend(candidate for candidate in data_dir.rglob("*") if candidate.is_file())
            candidate_paths.extend(
                candidate
                for candidate in source_dir.iterdir()
                if candidate.is_file() and candidate.name != "fetch-result.json" and not candidate.name.startswith(".")
            )

            if not candidate_paths:
                _write_failure(
                    failure_writer,
                    source.source_id,
                    str(source_dir),
                    "source_scan",
                    "scan",
                    "missing data directory",
                )
                source_failures += 1
            else:
                for path in sorted(candidate_paths, key=_file_sort_key):
                    if path in handled_paths:
                        continue

                    lower_suffix = path.suffix.lower()
                    if not lower_suffix and path.with_suffix(".hdr").exists():
                        handled_paths.add(path)
                        continue

                    try:
                        if lower_suffix == ".hdr":
                            records = _iter_envi_spectra(path, source.source_id, source.name, source.ingest_role)
                            handled_paths.add(path)
                            companion_path = _resolve_envi_data_path(path, _parse_envi_header(path))
                            handled_paths.add(companion_path)
                        elif lower_suffix in {".txt", ".csv", ".tab"}:
                            lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
                            records = _iter_textual_spectra_from_lines(
                                lines,
                                path,
                                source.source_id,
                                source.name,
                                source.ingest_role,
                            )
                            handled_paths.add(path)
                        elif lower_suffix == ".zip":
                            records = _iter_zip_spectra(path, source.source_id, source.name, source.ingest_role)
                            handled_paths.add(path)
                        elif lower_suffix in {".xlsx", ".xls"}:
                            records = _iter_xlsx_spectra(path, source.source_id, source.name, source.ingest_role)
                            handled_paths.add(path)
                        elif lower_suffix == ".rds":
                            records = _iter_rds_spectra(path, source.source_id, source.name, source.ingest_role)
                            handled_paths.add(path)
                        elif lower_suffix == ".nc":
                            records = _iter_netcdf_spectra(path, source.source_id, source.name, source.ingest_role)
                            handled_paths.add(path)
                        elif lower_suffix in {".json", ".pdf", ".docx", ".xml", ".sli", ".jpg", ".jpeg", ".png", ".md"} or not lower_suffix:
                            if lower_suffix != ".sli":
                                _write_failure(
                                    failure_writer,
                                    source.source_id,
                                    str(path),
                                    "file_dispatch",
                                    "dispatch",
                                    f"unsupported file type {lower_suffix or '[no suffix]'}",
                                )
                                source_failures += 1
                            handled_paths.add(path)
                            continue
                        else:
                            _write_failure(
                                failure_writer,
                                source.source_id,
                                str(path),
                                "file_dispatch",
                                "dispatch",
                                f"unsupported file type {lower_suffix}",
                            )
                            source_failures += 1
                            handled_paths.add(path)
                            continue
                    except Exception as exc:
                        _write_failure(
                            failure_writer,
                            source.source_id,
                            str(path),
                            "file_parse",
                            "parse",
                            str(exc),
                        )
                        source_failures += 1
                        handled_paths.add(path)
                        continue

                    try:
                        record_iterator = iter(records)
                        first_record = next(record_iterator, None)
                    except Exception as exc:
                        _write_failure(
                            failure_writer,
                            source.source_id,
                            str(path),
                            "file_parse",
                            "parse",
                            str(exc),
                        )
                        source_failures += 1
                        continue

                    if first_record is None:
                        _write_failure(
                            failure_writer,
                            source.source_id,
                            str(path),
                            "file_parse",
                            "parse",
                            "no spectra detected",
                        )
                        source_failures += 1
                        continue

                    for record in itertools.chain((first_record,), record_iterator):
                        try:
                            cleaned_wavelengths, cleaned_values, scale = _clean_spectrum(record)
                            normalized_values = _interpolate_to_grid(cleaned_wavelengths, cleaned_values)
                        except Exception as exc:
                            _write_failure(
                                failure_writer,
                                source.source_id,
                                record.input_path,
                                record.parser,
                                "normalize",
                                str(exc),
                            )
                            source_failures += 1
                            continue

                        spectrum_index += 1
                        spectrum_id = f"{source.source_id}_{spectrum_index:06d}"
                        normalized_count = sum(1 for value in normalized_values if value is not None)
                        metadata_writer.writerow(
                            [
                                source.source_id,
                                source.name,
                                source.ingest_role,
                                spectrum_id,
                                record.sample_name,
                                record.input_path,
                                record.parser,
                                len(cleaned_wavelengths),
                                min(cleaned_wavelengths),
                                max(cleaned_wavelengths),
                                _median_spacing_nm(cleaned_wavelengths) or "",
                                scale,
                                normalized_count,
                                json.dumps(record.metadata, sort_keys=True),
                            ]
                        )
                        normalized_writer.writerow(
                            [
                                source.source_id,
                                spectrum_id,
                                record.sample_name,
                                *["" if value is None else value for value in normalized_values],
                            ]
                        )
                        parser_counts[record.parser] += 1
                        source_parsers[record.parser] += 1
                        total_spectra += 1
                        source_spectra += 1

            summary_writer.writerow(
                [
                    source.source_id,
                    source.name,
                    source.ingest_role,
                    source_spectra,
                    source_failures,
                    " | ".join(f"{parser}:{count}" for parser, count in sorted(source_parsers.items())),
                ]
            )
            if source_spectra > 0:
                normalized_sources += 1
            total_failures += source_failures
    finally:
        metadata_handle.close()
        normalized_handle.close()
        failure_handle.close()
        summary_handle.close()

    _build_database(output_root)

    summary = {
        "selected_sources": len(selected),
        "downloaded_sources": downloaded_sources,
        "normalized_sources": normalized_sources,
        "normalized_spectra": total_spectra,
        "failure_rows": total_failures,
        "target_start_nm": TARGET_START_NM,
        "target_end_nm": TARGET_END_NM,
        "target_step_nm": TARGET_STEP_NM,
        "parser_counts": dict(sorted(parser_counts.items())),
    }
    summary_path = output_root / "build_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary
