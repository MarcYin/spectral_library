from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from urllib.request import urlopen

import numpy as np
try:
    import matplotlib
    from openpyxl import load_workbook

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except ModuleNotFoundError as exc:
    missing_name = exc.name or "an optional dependency"
    raise SystemExit(
        "scripts/build_official_mapping_examples.py requires the optional "
        f"internal-build dependencies (missing {missing_name}). Install them "
        'with `python3 -m pip install ".[internal-build]"`.'
    ) from exc


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = REPO_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from spectral_library import SpectralMapper


CANONICAL_WAVELENGTHS = np.arange(400, 2501, dtype=np.float64)
SEMANTIC_BANDS = ("ultra_blue", "blue", "green", "red", "nir", "swir1", "swir2")
EXAMPLES_ROOT = REPO_ROOT / "examples" / "official_mapping"
SRF_ROOT = EXAMPLES_ROOT / "srfs"
SIAC_ROOT = EXAMPLES_ROOT / "siac"
QUERIES_ROOT = EXAMPLES_ROOT / "queries"
RESULTS_ROOT = EXAMPLES_ROOT / "results"
DOCS_ASSETS_ROOT = REPO_ROOT / "docs" / "assets"
RETRIEVAL_DATE = "2026-03-15"
TMP_ROOT = REPO_ROOT / "tmp"


@dataclass(frozen=True)
class BandSelection:
    band_id: str
    official_band: str
    segment: str
    source_key: str
    description: str


@dataclass(frozen=True)
class SensorSelection:
    sensor_id: str
    display_name: str
    short_label: str
    filename: str
    source_agency: str
    source_document: str
    source_url: str
    source_note: str
    bands: tuple[BandSelection, ...]


OFFICIAL_SENSORS = (
    SensorSelection(
        sensor_id="modis_terra",
        display_name="Terra MODIS",
        short_label="MODIS Terra",
        filename="modis_terra.json",
        source_agency="NASA MCST",
        source_document="Terra_RSR_in-band.xlsx",
        source_url="https://mcst.gsfc.nasa.gov/sites/default/files/file_attachments/Terra_RSR_in-band.xlsx",
        source_note="Official Terra MODIS in-band relative spectral response workbook.",
        bands=(
            BandSelection("blue", "Band 3", "vnir", "band_3", "Blue"),
            BandSelection("green", "Band 4", "vnir", "band_4", "Green"),
            BandSelection("red", "Band 1", "vnir", "band_1", "Red"),
            BandSelection("nir", "Band 2", "vnir", "band_2", "Near infrared"),
            BandSelection("swir1", "Band 6", "swir", "band_6", "Shortwave infrared 1"),
            BandSelection("swir2", "Band 7", "swir", "band_7", "Shortwave infrared 2"),
        ),
    ),
    SensorSelection(
        sensor_id="sentinel2a_msi",
        display_name="Sentinel-2A MSI",
        short_label="Sentinel-2A",
        filename="sentinel2a_msi.json",
        source_agency="ESA Copernicus",
        source_document="COPE-GSEG-EOPG-TN-15-0007 - Sentinel-2 Spectral Response Functions 2024 - 4.0.xlsx",
        source_url="https://sentiwiki.copernicus.eu/__attachments/1692737/COPE-GSEG-EOPG-TN-15-0007%20-%20Sentinel-2%20Spectral%20Response%20Functions%202024%20-%204.0.xlsx",
        source_note="Official Sentinel-2 spectral response workbook; the examples use the S2A sheet.",
        bands=(
            BandSelection("ultra_blue", "B1", "vnir", "S2A_SR_AV_B1", "Coastal aerosol / ultra blue"),
            BandSelection("blue", "B2", "vnir", "S2A_SR_AV_B2", "Blue"),
            BandSelection("green", "B3", "vnir", "S2A_SR_AV_B3", "Green"),
            BandSelection("red", "B4", "vnir", "S2A_SR_AV_B4", "Red"),
            BandSelection("nir", "B8", "vnir", "S2A_SR_AV_B8", "Near infrared"),
            BandSelection("swir1", "B11", "swir", "S2A_SR_AV_B11", "Shortwave infrared 1"),
            BandSelection("swir2", "B12", "swir", "S2A_SR_AV_B12", "Shortwave infrared 2"),
        ),
    ),
    SensorSelection(
        sensor_id="landsat8_oli",
        display_name="Landsat 8 OLI",
        short_label="Landsat 8",
        filename="landsat8_oli.json",
        source_agency="USGS",
        source_document="Spectral Characteristics Viewer band JSON",
        source_url="https://landsat.usgs.gov/spectral-characteristics-viewer",
        source_note="Official USGS Spectral Characteristics Viewer JSON for Landsat 8 OLI/TIRS reflective bands 1-7.",
        bands=(
            BandSelection("ultra_blue", "Band 1", "vnir", "Landsat8OLITIRS1.json", "Coastal aerosol / ultra blue"),
            BandSelection("blue", "Band 2", "vnir", "Landsat8OLITIRS2.json", "Blue"),
            BandSelection("green", "Band 3", "vnir", "Landsat8OLITIRS3.json", "Green"),
            BandSelection("red", "Band 4", "vnir", "Landsat8OLITIRS4.json", "Red"),
            BandSelection("nir", "Band 5", "vnir", "Landsat8OLITIRS5.json", "Near infrared"),
            BandSelection("swir1", "Band 6", "swir", "Landsat8OLITIRS6.json", "Shortwave infrared 1"),
            BandSelection("swir2", "Band 7", "swir", "Landsat8OLITIRS7.json", "Shortwave infrared 2"),
        ),
    ),
    SensorSelection(
        sensor_id="landsat9_oli",
        display_name="Landsat 9 OLI",
        short_label="Landsat 9",
        filename="landsat9_oli.json",
        source_agency="USGS",
        source_document="Spectral Characteristics Viewer band JSON",
        source_url="https://landsat.usgs.gov/spectral-characteristics-viewer",
        source_note="Official USGS Spectral Characteristics Viewer JSON for Landsat 9 OLI/TIRS reflective bands 1-7.",
        bands=(
            BandSelection("ultra_blue", "Band 1", "vnir", "Landsat9OLITIRS1.json", "Coastal aerosol / ultra blue"),
            BandSelection("blue", "Band 2", "vnir", "Landsat9OLITIRS2.json", "Blue"),
            BandSelection("green", "Band 3", "vnir", "Landsat9OLITIRS3.json", "Green"),
            BandSelection("red", "Band 4", "vnir", "Landsat9OLITIRS4.json", "Red"),
            BandSelection("nir", "Band 5", "vnir", "Landsat9OLITIRS5.json", "Near infrared"),
            BandSelection("swir1", "Band 6", "swir", "Landsat9OLITIRS6.json", "Shortwave infrared 1"),
            BandSelection("swir2", "Band 7", "swir", "Landsat9OLITIRS7.json", "Shortwave infrared 2"),
        ),
    ),
)

SENSOR_BY_ID = {sensor.sensor_id: sensor for sensor in OFFICIAL_SENSORS}
PLOT_COLORS = {
    "modis_terra": "#0b5d8b",
    "sentinel2a_msi": "#e76f51",
    "landsat8_oli": "#2a9d8f",
    "landsat9_oli": "#6d597a",
}
PLOT_WINDOWS = {
    "ultra_blue": (410, 470),
    "blue": (430, 520),
    "green": (515, 595),
    "red": (620, 700),
    "nir": (810, 900),
    "swir1": (1540, 1685),
    "swir2": (2060, 2240),
}
PAIRWISE_SOURCE_TARGETS = (
    ("modis_terra", "sentinel2a_msi"),
    ("sentinel2a_msi", "landsat9_oli"),
    ("landsat8_oli", "modis_terra"),
)


def _download(url: str, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with urlopen(url, timeout=60) as response:
        destination.write_bytes(response.read())
    return destination


def _json_write(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=False) + "\n", encoding="utf-8")


def _weighted_center(wavelength_nm: np.ndarray, rsr: np.ndarray) -> float:
    return float(np.dot(wavelength_nm, rsr) / rsr.sum())


def _fwhm_nm(wavelength_nm: np.ndarray, rsr: np.ndarray) -> float:
    mask = rsr >= 0.5
    if not np.any(mask):
        return float(wavelength_nm[-1] - wavelength_nm[0])
    return float(wavelength_nm[mask][-1] - wavelength_nm[mask][0])


def _band_payload(band_id: str, segment: str, wavelengths_nm: np.ndarray, rsr: np.ndarray) -> dict[str, object]:
    support = rsr > 0
    return {
        "band_id": band_id,
        "segment": segment,
        "wavelength_nm": [round(float(value), 4) for value in wavelengths_nm.tolist()],
        "rsr": [round(float(value), 8) for value in rsr.tolist()],
        "center_nm": round(_weighted_center(wavelengths_nm, rsr), 4),
        "fwhm_nm": round(_fwhm_nm(wavelengths_nm, rsr), 4),
        "support_min_nm": round(float(wavelengths_nm[support][0]), 4),
        "support_max_nm": round(float(wavelengths_nm[support][-1]), 4),
    }


def _parse_modis_sensor(selection: SensorSelection, workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    bands: list[dict[str, object]] = []
    for band in selection.bands:
        worksheet = workbook[band.source_key]
        samples_by_wavelength: dict[float, list[float]] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or row[2] is None or row[3] is None:
                continue
            wavelength = float(row[2])
            value = float(row[3])
            if value <= 0:
                continue
            samples_by_wavelength.setdefault(wavelength, []).append(value)
        wavelengths_nm = np.asarray(sorted(samples_by_wavelength), dtype=np.float64)
        rsr = np.asarray(
            [float(np.mean(samples_by_wavelength[wavelength])) for wavelength in wavelengths_nm],
            dtype=np.float64,
        )
        bands.append(_band_payload(band.band_id, band.segment, wavelengths_nm, rsr))
    return {"sensor_id": selection.sensor_id, "bands": bands}


def _parse_sentinel_sensor(selection: SensorSelection, workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["Spectral Responses (S2A)"]
    rows = list(worksheet.iter_rows(values_only=True))
    header = [value if value is not None else "" for value in rows[0]]
    columns = {str(name): index for index, name in enumerate(header)}
    wavelength_index = columns["SR_WL"]
    bands: list[dict[str, object]] = []
    for band in selection.bands:
        response_index = columns[band.source_key]
        wavelengths: list[float] = []
        response: list[float] = []
        for row in rows[1:]:
            if row[wavelength_index] is None or row[response_index] is None:
                continue
            rsr = float(row[response_index])
            if rsr <= 0:
                continue
            wavelengths.append(float(row[wavelength_index]))
            response.append(rsr)
        bands.append(
            _band_payload(
                band.band_id,
                band.segment,
                np.asarray(wavelengths, dtype=np.float64),
                np.asarray(response, dtype=np.float64),
            )
        )
    return {"sensor_id": selection.sensor_id, "bands": bands}


def _parse_landsat_sensor(selection: SensorSelection) -> dict[str, object]:
    base_url = "https://landsat.usgs.gov/landsat/spectral_viewer/c3-master/htdocs/data/bands"
    bands: list[dict[str, object]] = []
    for band in selection.bands:
        with urlopen(f"{base_url}/{band.source_key}", timeout=60) as response:
            payload = json.loads(response.read().decode("utf-8"))
        wavelength_key, response_key = list(payload[0])
        wavelengths_nm = np.asarray([1000.0 * float(item[wavelength_key]) for item in payload], dtype=np.float64)
        rsr = np.asarray([float(item[response_key]) for item in payload], dtype=np.float64)
        positive = rsr > 0
        bands.append(_band_payload(band.band_id, band.segment, wavelengths_nm[positive], rsr[positive]))
    return {"sensor_id": selection.sensor_id, "bands": bands}


def _sigmoid(center_nm: float, scale_nm: float) -> np.ndarray:
    return 1.0 / (1.0 + np.exp(-(CANONICAL_WAVELENGTHS - center_nm) / scale_nm))


def _gaussian(center_nm: float, width_nm: float) -> np.ndarray:
    return np.exp(-0.5 * ((CANONICAL_WAVELENGTHS - center_nm) / width_nm) ** 2)


def _clip_reflectance(values: np.ndarray) -> np.ndarray:
    return np.clip(values, 0.0, 0.95)


def build_library_spectra() -> dict[str, np.ndarray]:
    w = CANONICAL_WAVELENGTHS
    water_abs = 0.16 * _gaussian(1400, 55) + 0.23 * _gaussian(1900, 85) + 0.05 * _gaussian(2200, 55)

    dense_vegetation = _clip_reflectance(
        0.035
        + 0.055 * _gaussian(550, 30)
        - 0.03 * _gaussian(670, 22)
        + 0.53 * _sigmoid(720, 16)
        - 0.00022 * np.clip(w - 1100, 0, None)
        - water_abs
    )
    dry_grass = _clip_reflectance(
        0.08
        + 0.05 * _gaussian(560, 42)
        - 0.02 * _gaussian(675, 28)
        + 0.28 * _sigmoid(725, 22)
        - 0.00012 * np.clip(w - 1200, 0, None)
        - 0.11 * _gaussian(1450, 70)
        - 0.12 * _gaussian(1940, 95)
        - 0.03 * _gaussian(2190, 65)
    )
    shrubland = _clip_reflectance(
        0.06
        + 0.045 * _gaussian(555, 35)
        - 0.018 * _gaussian(668, 24)
        + 0.38 * _sigmoid(718, 18)
        - 0.00015 * np.clip(w - 1180, 0, None)
        - 0.13 * _gaussian(1430, 62)
        - 0.14 * _gaussian(1925, 90)
        - 0.035 * _gaussian(2210, 60)
    )
    bright_soil = _clip_reflectance(
        0.16
        + 0.00012 * (w - 400)
        + 0.028 * _gaussian(1650, 140)
        - 0.05 * _gaussian(1410, 60)
        - 0.06 * _gaussian(1915, 95)
        - 0.035 * _gaussian(2215, 70)
    )
    dark_soil = _clip_reflectance(
        0.09
        + 0.00008 * (w - 400)
        + 0.015 * _gaussian(1650, 150)
        - 0.03 * _gaussian(1410, 70)
        - 0.04 * _gaussian(1910, 95)
        - 0.025 * _gaussian(2210, 80)
    )
    concrete = _clip_reflectance(
        0.17
        + 0.00007 * (w - 400)
        + 0.02 * _gaussian(600, 95)
        - 0.012 * _gaussian(1380, 90)
        - 0.02 * _gaussian(1920, 100)
        - 0.014 * _gaussian(2210, 85)
    )
    asphalt = _clip_reflectance(
        0.05
        + 0.00005 * (w - 400)
        + 0.01 * _gaussian(900, 180)
        - 0.01 * _gaussian(1410, 80)
        - 0.015 * _gaussian(1920, 100)
    )
    clear_water = _clip_reflectance(
        0.04
        - 0.000045 * (w - 400)
        + 0.01 * _gaussian(480, 35)
        - 0.02 * _sigmoid(760, 25)
        - 0.015 * _sigmoid(1000, 25)
    )
    turbid_water = _clip_reflectance(
        0.08
        - 0.00005 * (w - 400)
        + 0.03 * _gaussian(620, 90)
        - 0.03 * _sigmoid(780, 25)
        - 0.02 * _sigmoid(1050, 30)
    )
    gypsum_sand = _clip_reflectance(
        0.24
        + 0.00008 * (w - 400)
        + 0.018 * _gaussian(1450, 120)
        - 0.025 * _gaussian(1750, 80)
        - 0.03 * _gaussian(2210, 75)
    )

    return {
        "dense_vegetation": dense_vegetation,
        "dry_grass": dry_grass,
        "shrubland": shrubland,
        "bright_soil": bright_soil,
        "dark_soil": dark_soil,
        "concrete": concrete,
        "asphalt": asphalt,
        "clear_water": clear_water,
        "turbid_water": turbid_water,
        "gypsum_sand": gypsum_sand,
    }


def build_truth_spectra(library_spectra: dict[str, np.ndarray]) -> dict[str, np.ndarray]:
    return {
        "veg_soil_mix": _clip_reflectance(0.68 * library_spectra["dense_vegetation"] + 0.32 * library_spectra["bright_soil"]),
        "urban_soil_mix": _clip_reflectance(0.55 * library_spectra["concrete"] + 0.45 * library_spectra["dark_soil"]),
        "water_edge_mix": _clip_reflectance(0.58 * library_spectra["turbid_water"] + 0.42 * library_spectra["dry_grass"]),
    }


def _write_siac_fixture(library_spectra: dict[str, np.ndarray]) -> None:
    tabular_root = SIAC_ROOT / "tabular"
    tabular_root.mkdir(parents=True, exist_ok=True)
    metadata_path = tabular_root / "siac_spectra_metadata.csv"
    spectra_path = tabular_root / "siac_normalized_spectra.csv"

    metadata_rows = [
        {
            "source_id": "official_example",
            "spectrum_id": sample_name,
            "sample_name": sample_name,
            "source_name": "Official sensor example synthetic library",
            "landcover_group": sample_name,
        }
        for sample_name in library_spectra
    ]
    with metadata_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["source_id", "spectrum_id", "sample_name", "source_name", "landcover_group"],
        )
        writer.writeheader()
        writer.writerows(metadata_rows)

    nm_columns = [f"nm_{int(wavelength)}" for wavelength in CANONICAL_WAVELENGTHS]
    with spectra_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["source_id", "spectrum_id", "sample_name", *nm_columns])
        writer.writeheader()
        for sample_name, spectrum in library_spectra.items():
            row = {
                "source_id": "official_example",
                "spectrum_id": sample_name,
                "sample_name": sample_name,
            }
            row.update({column: round(float(value), 8) for column, value in zip(nm_columns, spectrum)})
            writer.writerow(row)


def _simulate_sensor(sensor_schema: dict[str, object], spectrum: np.ndarray) -> dict[str, float]:
    values: dict[str, float] = {}
    for band in sensor_schema["bands"]:  # type: ignore[index]
        wavelengths = np.asarray(band["wavelength_nm"], dtype=np.float64)  # type: ignore[index]
        response = np.asarray(band["rsr"], dtype=np.float64)  # type: ignore[index]
        sampled = np.interp(wavelengths, CANONICAL_WAVELENGTHS, spectrum)
        values[str(band["band_id"])] = float(np.dot(sampled, response) / response.sum())
    return values


def _write_single_query(path: Path, reflectance: dict[str, float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["band_id", "reflectance"])
        writer.writeheader()
        for band_id, value in reflectance.items():
            writer.writerow({"band_id": band_id, "reflectance": round(float(value), 8)})


def _write_batch_query(path: Path, sample_rows: dict[str, dict[str, float]], band_ids: tuple[str, ...]) -> None:
    fieldnames = ["sample_id", *band_ids]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for sample_id, values in sample_rows.items():
            row = {"sample_id": sample_id}
            row.update({band_id: round(float(values[band_id]), 8) for band_id in band_ids})
            writer.writerow(row)


def _python_env() -> dict[str, str]:
    env = dict(os.environ)
    existing = env.get("PYTHONPATH")
    env["PYTHONPATH"] = str(SRC_ROOT) if not existing else str(SRC_ROOT) + os.pathsep + existing
    return env


def _run_cli(*args: str) -> str:
    command = [sys.executable, "-m", "spectral_library.cli", *args]
    completed = subprocess.run(
        command,
        check=False,
        cwd=REPO_ROOT,
        env=_python_env(),
        text=True,
        capture_output=True,
    )
    if completed.returncode != 0:
        if completed.stdout:
            print(completed.stdout, file=sys.stdout)
        if completed.stderr:
            print(completed.stderr, file=sys.stderr)
        raise subprocess.CalledProcessError(completed.returncode, command, output=completed.stdout, stderr=completed.stderr)
    return completed.stdout


def _prepare_example_runtime(prepared_root: Path) -> None:
    if prepared_root.exists():
        shutil.rmtree(prepared_root)
    _run_cli(
        "prepare-mapping-library",
        "--siac-root",
        str(SIAC_ROOT),
        "--srf-root",
        str(SRF_ROOT),
        "--source-sensor",
        "modis_terra",
        "--source-sensor",
        "sentinel2a_msi",
        "--source-sensor",
        "landsat8_oli",
        "--source-sensor",
        "landsat9_oli",
        "--output-root",
        str(prepared_root),
    )


def _write_truth_tables(truth_spectra: dict[str, np.ndarray], sensors: dict[str, dict[str, object]]) -> None:
    truth_root = RESULTS_ROOT / "truth"
    truth_root.mkdir(parents=True, exist_ok=True)
    for sensor_id, schema in sensors.items():
        band_ids = tuple(str(band["band_id"]) for band in schema["bands"])  # type: ignore[index]
        path = truth_root / f"{sensor_id}_truth.csv"
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["sample_id", *band_ids])
            writer.writeheader()
            for sample_id, spectrum in truth_spectra.items():
                row = {"sample_id": sample_id}
                row.update(
                    {
                        band_id: round(float(value), 8)
                        for band_id, value in _simulate_sensor(schema, spectrum).items()
                    }
                )
                writer.writerow(row)


def _collect_reflectance_rows(truth_spectra: dict[str, np.ndarray], sensors: dict[str, dict[str, object]]) -> dict[str, dict[str, dict[str, float]]]:
    by_sensor: dict[str, dict[str, dict[str, float]]] = {}
    for sensor_id, schema in sensors.items():
        sensor_rows: dict[str, dict[str, float]] = {}
        for sample_id, spectrum in truth_spectra.items():
            sensor_rows[sample_id] = _simulate_sensor(schema, spectrum)
        by_sensor[sensor_id] = sensor_rows
    return by_sensor


def _load_result_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _mean_abs_error(mapped: dict[str, float], truth: dict[str, float]) -> float:
    common = [band_id for band_id in truth if band_id in mapped]
    values = [abs(mapped[band_id] - truth[band_id]) for band_id in common]
    return float(np.mean(values))


def _rmse(mapped: dict[str, float], truth: dict[str, float]) -> float:
    common = [band_id for band_id in truth if band_id in mapped]
    values = [(mapped[band_id] - truth[band_id]) ** 2 for band_id in common]
    return float(np.sqrt(np.mean(values)))


def _write_selected_cli_outputs() -> None:
    selected_root = RESULTS_ROOT / "selected"
    selected_root.mkdir(parents=True, exist_ok=True)

    query_map = {
        "modis_terra": QUERIES_ROOT / "single" / "veg_soil_mix_modis_terra.csv",
        "sentinel2a_msi": QUERIES_ROOT / "single" / "veg_soil_mix_sentinel2a_msi.csv",
        "landsat8_oli": QUERIES_ROOT / "single" / "veg_soil_mix_landsat8_oli.csv",
    }

    with tempfile.TemporaryDirectory(prefix="official_mapping_runtime_", dir=str(TMP_ROOT)) as tmpdir:
        prepared_root = Path(tmpdir) / "prepared"
        _prepare_example_runtime(prepared_root)

        _run_cli(
            "map-reflectance",
            "--prepared-root",
            str(prepared_root),
            "--source-sensor",
            "modis_terra",
            "--target-sensor",
            "sentinel2a_msi",
            "--input",
            str(query_map["modis_terra"]),
            "--output-mode",
            "target_sensor",
            "--k",
            "3",
            "--output",
            str(selected_root / "veg_soil_mix_modis_to_sentinel2a.csv"),
        )
        _run_cli(
            "map-reflectance",
            "--prepared-root",
            str(prepared_root),
            "--source-sensor",
            "sentinel2a_msi",
            "--target-sensor",
            "landsat9_oli",
            "--input",
            str(query_map["sentinel2a_msi"]),
            "--output-mode",
            "target_sensor",
            "--k",
            "3",
            "--output",
            str(selected_root / "veg_soil_mix_sentinel2a_to_landsat9.csv"),
        )
        _run_cli(
            "map-reflectance",
            "--prepared-root",
            str(prepared_root),
            "--source-sensor",
            "landsat8_oli",
            "--target-sensor",
            "modis_terra",
            "--input",
            str(query_map["landsat8_oli"]),
            "--output-mode",
            "target_sensor",
            "--k",
            "3",
            "--output",
            str(selected_root / "veg_soil_mix_landsat8_to_modis.csv"),
        )
        _run_cli(
            "map-reflectance",
            "--prepared-root",
            str(prepared_root),
            "--source-sensor",
            "sentinel2a_msi",
            "--input",
            str(query_map["sentinel2a_msi"]),
            "--output-mode",
            "full_spectrum",
            "--k",
            "3",
            "--output",
            str(selected_root / "veg_soil_mix_sentinel2a_full_spectrum.csv"),
        )
        _run_cli(
            "map-reflectance-batch",
            "--prepared-root",
            str(prepared_root),
            "--source-sensor",
            "landsat8_oli",
            "--target-sensor",
            "sentinel2a_msi",
            "--input",
            str(QUERIES_ROOT / "batch" / "landsat8_truth_batch.csv"),
            "--output-mode",
            "target_sensor",
            "--k",
            "3",
            "--output",
            str(selected_root / "landsat8_to_sentinel2a_batch.csv"),
            "--diagnostics-output",
            str(selected_root / "landsat8_to_sentinel2a_batch_diagnostics.json"),
        )


def _write_pairwise_metrics(truth_rows_by_sensor: dict[str, dict[str, dict[str, float]]]) -> dict[tuple[str, str], dict[str, float]]:
    metrics_root = RESULTS_ROOT / "metrics"
    metrics_root.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, object]] = []
    metrics: dict[tuple[str, str], dict[str, float]] = {}

    with tempfile.TemporaryDirectory(prefix="official_mapping_runtime_", dir=str(TMP_ROOT)) as tmpdir:
        prepared_root = Path(tmpdir) / "prepared"
        _prepare_example_runtime(prepared_root)
        mapper = SpectralMapper(prepared_root)

        for source_sensor in SENSOR_BY_ID:
            reflectance_rows = [truth_rows_by_sensor[source_sensor][sample_id] for sample_id in truth_rows_by_sensor[source_sensor]]
            sample_ids = tuple(truth_rows_by_sensor[source_sensor])
            for target_sensor in SENSOR_BY_ID:
                if source_sensor == target_sensor:
                    continue
                result = mapper.map_reflectance_batch(
                    source_sensor=source_sensor,
                    reflectance_rows=reflectance_rows,
                    sample_ids=sample_ids,
                    output_mode="target_sensor",
                    target_sensor=target_sensor,
                    k=3,
                )
                sample_mae: list[float] = []
                sample_rmse: list[float] = []
                for sample_id, mapping_result in zip(result.sample_ids, result.results):
                    if mapping_result.target_reflectance is None:
                        raise RuntimeError(
                            f"Missing target_reflectance for {source_sensor} -> {target_sensor} sample {sample_id}."
                        )
                    mapped = {
                        band_id: float(value)
                        for band_id, value in zip(mapping_result.target_band_ids, mapping_result.target_reflectance)
                    }
                    truth = truth_rows_by_sensor[target_sensor][sample_id]
                    sample_mae.append(_mean_abs_error(mapped, truth))
                    sample_rmse.append(_rmse(mapped, truth))

                payload = {
                    "mean_abs_error": float(np.mean(sample_mae)),
                    "rmse": float(np.mean(sample_rmse)),
                    "sample_count": len(sample_mae),
                }
                metrics[(source_sensor, target_sensor)] = payload
                rows.append(
                    {
                        "source_sensor": source_sensor,
                        "target_sensor": target_sensor,
                        "source_label": SENSOR_BY_ID[source_sensor].short_label,
                        "target_label": SENSOR_BY_ID[target_sensor].short_label,
                        **payload,
                    }
                )

    csv_path = metrics_root / "pairwise_band_metrics.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "source_sensor",
                "target_sensor",
                "source_label",
                "target_label",
                "mean_abs_error",
                "rmse",
                "sample_count",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    _json_write(metrics_root / "pairwise_band_metrics.json", {"pairs": rows})
    return metrics


def _plot_selected_bands(sensor_payloads: dict[str, dict[str, object]]) -> None:
    DOCS_ASSETS_ROOT.mkdir(parents=True, exist_ok=True)
    figure, axes = plt.subplots(4, 2, figsize=(14, 11), constrained_layout=True)
    blank_axis = axes.flatten()[-1]
    axes_by_band = {
        band_id: axes.flatten()[index]
        for index, band_id in enumerate(SEMANTIC_BANDS)
    }
    blank_axis.axis("off")

    for band_id, axis in axes_by_band.items():
        x_min, x_max = PLOT_WINDOWS[band_id]
        axis.set_title(band_id.replace("_", " ").title())
        axis.set_xlim(x_min, x_max)
        axis.set_ylim(0, 1.05)
        axis.set_xlabel("Wavelength (nm)")
        axis.set_ylabel("Relative response")
        axis.grid(alpha=0.25)
        has_any = False
        for sensor in OFFICIAL_SENSORS:
            payload = sensor_payloads[sensor.sensor_id]
            matching = [
                band
                for band in payload["bands"]  # type: ignore[index]
                if str(band["band_id"]) == band_id
            ]
            if not matching:
                continue
            has_any = True
            band = matching[0]
            axis.plot(
                np.asarray(band["wavelength_nm"], dtype=np.float64),
                np.asarray(band["rsr"], dtype=np.float64),
                label=sensor.short_label,
                linewidth=2.0,
                color=PLOT_COLORS[sensor.sensor_id],
            )
        if not has_any:
            axis.text(0.5, 0.5, "Not available", transform=axis.transAxes, ha="center", va="center")

    handles, labels = axes_by_band["blue"].get_legend_handles_labels()
    blank_axis.text(
        0.5,
        0.72,
        "Official selected band responses\nused by the example runtime",
        ha="center",
        va="center",
        fontsize=15,
    )
    blank_axis.legend(handles, labels, loc="center", ncol=1, frameon=False)
    figure.savefig(DOCS_ASSETS_ROOT / "official_sensor_selected_bands.png", dpi=180)
    plt.close(figure)


def _plot_pairwise_heatmap(metrics: dict[tuple[str, str], dict[str, float]]) -> None:
    order = [sensor.sensor_id for sensor in OFFICIAL_SENSORS]
    matrix = np.full((len(order), len(order)), np.nan, dtype=np.float64)
    for row_index, source_sensor in enumerate(order):
        for column_index, target_sensor in enumerate(order):
            if source_sensor == target_sensor:
                continue
            matrix[row_index, column_index] = metrics[(source_sensor, target_sensor)]["mean_abs_error"]

    figure, axis = plt.subplots(figsize=(7.5, 6))
    image = axis.imshow(matrix, cmap="YlOrRd")
    axis.set_xticks(range(len(order)), [SENSOR_BY_ID[sensor_id].short_label for sensor_id in order], rotation=20, ha="right")
    axis.set_yticks(range(len(order)), [SENSOR_BY_ID[sensor_id].short_label for sensor_id in order])
    axis.set_title("Mean absolute band error across three synthetic truth samples")
    for row_index in range(len(order)):
        for column_index in range(len(order)):
            if np.isnan(matrix[row_index, column_index]):
                axis.text(column_index, row_index, "-", ha="center", va="center", color="#555555")
            else:
                axis.text(
                    column_index,
                    row_index,
                    f"{matrix[row_index, column_index]:.4f}",
                    ha="center",
                    va="center",
                    color="#111111",
                )
    colorbar = figure.colorbar(image, ax=axis, shrink=0.86)
    colorbar.set_label("Mean absolute error (reflectance)")
    figure.tight_layout()
    figure.savefig(DOCS_ASSETS_ROOT / "official_pairwise_band_mae.png", dpi=180)
    plt.close(figure)


def _plot_modis_to_sentinel_example(sensor_payloads: dict[str, dict[str, object]], truth_rows_by_sensor: dict[str, dict[str, dict[str, float]]]) -> None:
    result_path = RESULTS_ROOT / "selected" / "veg_soil_mix_modis_to_sentinel2a.csv"
    rows = _load_result_rows(result_path)
    mapped = {row["band_id"]: float(row["reflectance"]) for row in rows}
    truth = truth_rows_by_sensor["sentinel2a_msi"]["veg_soil_mix"]
    band_ids = [band["band_id"] for band in sensor_payloads["sentinel2a_msi"]["bands"]]  # type: ignore[index]

    x = np.arange(len(band_ids))
    width = 0.38
    figure, axis = plt.subplots(figsize=(9.5, 4.8))
    axis.bar(x - width / 2, [truth[band_id] for band_id in band_ids], width=width, label="Truth", color="#2a9d8f")
    axis.bar(x + width / 2, [mapped[band_id] for band_id in band_ids], width=width, label="Mapped from MODIS", color="#e76f51")
    axis.set_xticks(x, [band_id.replace("_", "\n") for band_id in band_ids])
    axis.set_ylabel("Reflectance")
    axis.set_title("MODIS Terra to Sentinel-2A example on the veg_soil_mix query")
    axis.grid(axis="y", alpha=0.25)
    axis.legend(frameon=False)
    figure.tight_layout()
    figure.savefig(DOCS_ASSETS_ROOT / "official_modis_to_sentinel2a_example.png", dpi=180)
    plt.close(figure)


def _write_examples_readme() -> None:
    text = """# Official Sensor Mapping Example

This directory contains a fully reproducible mapping example built from:

- official Terra MODIS relative spectral response tables from NASA MCST,
- official Sentinel-2A MSI spectral response tables from ESA Copernicus,
- official Landsat 8 and Landsat 9 band JSON files from the USGS Spectral Characteristics Viewer.

Contents:

- `official_source_manifest.json`: upstream provenance for the selected band subsets
- `srfs/`: package-ready sensor JSON files for the selected bands
- `siac/`: a compact synthetic SIAC-style hyperspectral library on the public 400-2500 nm grid
- `queries/`: single-sample and batch query CSVs used in the docs
- `results/`: generated mapping outputs, truth tables, and pairwise error summaries

Regenerate everything from the official upstream sources with:

```bash
python3 scripts/build_official_mapping_examples.py
```

The public write-up that explains the example commands and figures is in
[`docs/official_sensor_examples.md`](../../docs/official_sensor_examples.md).
"""
    path = EXAMPLES_ROOT / "README.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def build_official_mapping_examples() -> None:
    TMP_ROOT.mkdir(parents=True, exist_ok=True)
    for root in (SRF_ROOT, QUERIES_ROOT, RESULTS_ROOT, DOCS_ASSETS_ROOT):
        root.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="official_sensor_sources_", dir=str(REPO_ROOT / "tmp")) as tmpdir:
        tmp_root = Path(tmpdir)
        modis_path = _download(SENSOR_BY_ID["modis_terra"].source_url, tmp_root / "Terra_RSR_in-band.xlsx")
        sentinel_path = _download(
            SENSOR_BY_ID["sentinel2a_msi"].source_url,
            tmp_root / "Sentinel-2_Spectral_Response_Functions_2024_4.0.xlsx",
        )

        sensor_payloads = {
            "modis_terra": _parse_modis_sensor(SENSOR_BY_ID["modis_terra"], modis_path),
            "sentinel2a_msi": _parse_sentinel_sensor(SENSOR_BY_ID["sentinel2a_msi"], sentinel_path),
            "landsat8_oli": _parse_landsat_sensor(SENSOR_BY_ID["landsat8_oli"]),
            "landsat9_oli": _parse_landsat_sensor(SENSOR_BY_ID["landsat9_oli"]),
        }

    for sensor in OFFICIAL_SENSORS:
        _json_write(SRF_ROOT / sensor.filename, sensor_payloads[sensor.sensor_id])

    provenance_payload = {
        "retrieved_on": RETRIEVAL_DATE,
        "description": "Official-source band subsets used for the documentation and mapping examples.",
        "sensors": [
            {
                "sensor_id": sensor.sensor_id,
                "display_name": sensor.display_name,
                "source_agency": sensor.source_agency,
                "source_document": sensor.source_document,
                "source_url": sensor.source_url,
                "source_note": sensor.source_note,
                "selected_bands": [
                    {
                        "band_id": band.band_id,
                        "official_band": band.official_band,
                        "segment": band.segment,
                        "description": band.description,
                    }
                    for band in sensor.bands
                ],
            }
            for sensor in OFFICIAL_SENSORS
        ],
    }
    _json_write(EXAMPLES_ROOT / "official_source_manifest.json", provenance_payload)

    library_spectra = build_library_spectra()
    truth_spectra = build_truth_spectra(library_spectra)
    _write_siac_fixture(library_spectra)

    truth_rows_by_sensor = _collect_reflectance_rows(truth_spectra, sensor_payloads)
    for sensor_id, sample_rows in truth_rows_by_sensor.items():
        _write_single_query(QUERIES_ROOT / "single" / f"veg_soil_mix_{sensor_id}.csv", sample_rows["veg_soil_mix"])

    landsat8_band_ids = tuple(str(band["band_id"]) for band in sensor_payloads["landsat8_oli"]["bands"])  # type: ignore[index]
    _write_batch_query(QUERIES_ROOT / "batch" / "landsat8_truth_batch.csv", truth_rows_by_sensor["landsat8_oli"], landsat8_band_ids)
    _write_truth_tables(truth_spectra, sensor_payloads)
    _write_selected_cli_outputs()
    metrics = _write_pairwise_metrics(truth_rows_by_sensor)
    _plot_selected_bands(sensor_payloads)
    _plot_pairwise_heatmap(metrics)
    _plot_modis_to_sentinel_example(sensor_payloads, truth_rows_by_sensor)
    _write_examples_readme()


def main() -> int:
    build_official_mapping_examples()
    print(json.dumps({"status": "ok", "examples_root": str(EXAMPLES_ROOT)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
